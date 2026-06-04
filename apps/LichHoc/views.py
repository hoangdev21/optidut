from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from django.db.models import Count, Q, Case, When, Value, IntegerField
from django.utils import timezone
from datetime import datetime, timedelta
import io
import csv
import json


from .models import LichHoc, LopHoc, MonHoc, DangKyHocPhan, YeuCauDoiLich, KHUNG_GIO_TIET
from .forms import FormLichHoc, FormLopHoc
from apps.PhongHoc.models import PhongHoc


def cap_nhat_trang_thai_phong(ngay=None):
    """Cập nhật trạng thái phòng dựa trên lịch học."""
    if ngay is None:
        ngay = timezone.now().date()
    # Phòng đang có lịch hoạt động hôm nay
    phong_dang_dung_ids = LichHoc.objects.filter(
        ngay_hoc=ngay, trang_thai='hoat_dong'
    ).values_list('phong_hoc_id', flat=True).distinct()
    # Cập nhật phòng đang sử dụng
    PhongHoc.objects.filter(id__in=phong_dang_dung_ids).exclude(
        trang_thai='bao_tri'
    ).update(trang_thai='dang_su_dung')
    # Cập nhật phòng trống (không có lịch hôm nay và không bảo trì)
    PhongHoc.objects.exclude(id__in=phong_dang_dung_ids).exclude(
        trang_thai='bao_tri'
    ).update(trang_thai='trong')


def kiem_tra_quyen_lich(view_func):
    """Decorator: chỉ giáo vụ mới được quản lý lịch học."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('dang_nhap')
        if not (request.user.la_giao_vu or request.user.la_quan_tri):
            messages.error(request, 'Chỉ giáo vụ hoặc quản trị viên mới có quyền quản lý lịch học.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@login_required
def danh_sach_lich(request):
    """Danh sách lịch học - mặc định hôm nay, có nút chuyển ngày."""
    hom_nay = timezone.now().date()
    
    # Cập nhật trạng thái phòng
    cap_nhat_trang_thai_phong(hom_nay)
    
    # Lọc theo ngày — mặc định hôm nay
    ngay_loc = request.GET.get('ngay', '')
    if ngay_loc:
        try:
            ngay_hien_tai = datetime.strptime(ngay_loc, '%Y-%m-%d').date()
        except ValueError:
            ngay_hien_tai = hom_nay
    else:
        ngay_hien_tai = hom_nay
    
    ngay_truoc = ngay_hien_tai - timedelta(days=1)
    ngay_sau = ngay_hien_tai + timedelta(days=1)
    
    # Khởi tạo base queryset với các bộ lọc tìm kiếm và nâng cao (chung cho cả ngày, tuần, năm)
    base_qs = LichHoc.objects.select_related(
        'mon_hoc', 'giang_vien', 'phong_hoc', 'lop_hoc'
    )

    # PHÂN QUYỀN HIỂN THỊ DỮ LIỆU
    lop_ids = None
    if request.user.la_sinh_vien:
        lop_ids = DangKyHocPhan.objects.filter(sinh_vien=request.user).values_list('lop_hoc_id', flat=True)
        base_qs = base_qs.filter(lop_hoc_id__in=lop_ids)
    elif request.user.la_giang_vien:
        base_qs = base_qs.filter(giang_vien=request.user)

    # Tìm kiếm
    tu_khoa = request.GET.get('q', '')
    if tu_khoa:
        base_qs = base_qs.filter(
            Q(mon_hoc__ten_mon__icontains=tu_khoa) |
            Q(mon_hoc__ma_mon__icontains=tu_khoa)
        )

    # Bộ lọc nâng cao
    phong_loc = request.GET.get('phong', '')
    if phong_loc:
        if phong_loc.isdigit():
            base_qs = base_qs.filter(phong_hoc_id=phong_loc)
        else:
            base_qs = base_qs.filter(phong_hoc__ma_phong=phong_loc)

    lop_loc = request.GET.get('lop', '')
    if lop_loc:
        base_qs = base_qs.filter(lop_hoc_id=lop_loc)

    buoi_loc = request.GET.get('buoi', '')
    buoi_map = {
        'sang': (1, 5),
        'chieu': (6, 10),
        'toi': (11, 14),
    }
    if buoi_loc in buoi_map:
        bd, kt = buoi_map[buoi_loc]
        base_qs = base_qs.filter(tiet_bat_dau__lte=kt, tiet_ket_thuc__gte=bd)

    tiet_loc = request.GET.get('tiet', '')
    if tiet_loc:
        try:
            tiet_loc_int = int(tiet_loc)
            base_qs = base_qs.filter(tiet_bat_dau__lte=tiet_loc_int, tiet_ket_thuc__gte=tiet_loc_int)
        except ValueError:
            tiet_loc = ''

    trang_thai = request.GET.get('trang_thai', '')
    if trang_thai:
        base_qs = base_qs.filter(trang_thai=trang_thai)

    # 1. Queryset của ngày hiện tại (cho danh sách hiển thị mặc định và Table View)
    queryset = base_qs.filter(ngay_hoc=ngay_hien_tai)
    queryset = queryset.annotate(
        priority=Case(
            When(trang_thai='hoat_dong', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        )
    ).order_by('priority', 'tiet_bat_dau')

    # 2. Queryset của Tuần hiện tại (cho Week View)
    mon_date = ngay_hien_tai - timedelta(days=ngay_hien_tai.weekday())
    sun_date = mon_date + timedelta(days=6)
    week_queryset = base_qs.filter(ngay_hoc__range=[mon_date, sun_date], trang_thai='hoat_dong')
    
    # 3. Queryset của Năm hiện tại (cho Year View)
    year_queryset = base_qs.filter(ngay_hoc__year=ngay_hien_tai.year, trang_thai='hoat_dong')

    # Serialize Week Data
    week_data = []
    for item in week_queryset:
        week_data.append({
            'lop_hoc_id': item.lop_hoc.id if item.lop_hoc else None,
            'ten_lop': item.lop_hoc.ten_lop if item.lop_hoc else item.ma_lop,
            'mon_hoc_id': item.mon_hoc.id if item.mon_hoc else None,
            'ma_mon': item.mon_hoc.ma_mon if item.mon_hoc else '',
            'ten_mon': item.mon_hoc.ten_mon if item.mon_hoc else '',
            'giang_vien_id': item.giang_vien.id if item.giang_vien else None,
            'ma_giang_vien': item.giang_vien.ma_so if item.giang_vien else '',
            'ten_giang_vien': item.giang_vien.ho_ten if item.giang_vien else '',
            'phong_hoc_id': item.phong_hoc.id if item.phong_hoc else None,
            'ma_phong': item.phong_hoc.ma_phong if item.phong_hoc else '',
            'ngay_hoc': item.ngay_hoc.strftime('%Y-%m-%d'),
            'tiet_bat_dau': item.tiet_bat_dau,
            'tiet_ket_thuc': item.tiet_ket_thuc,
            'si_so': item.si_so,
            'ghi_chu': item.ghi_chu
        })

    # Group Year Data by Month
    year_counts_qs = year_queryset.values('ngay_hoc__month').annotate(count=Count('id'))
    year_counts = {item['ngay_hoc__month']: item['count'] for item in year_counts_qs}
    for m in range(1, 13):
        year_counts.setdefault(m, 0)

    # Serialize Table Data
    table_data = []
    for item in base_qs:
        table_data.append({
            'lop_hoc_id': item.lop_hoc.id if item.lop_hoc else None,
            'ten_lop': item.lop_hoc.ten_lop if item.lop_hoc else item.ma_lop,
            'mon_hoc_id': item.mon_hoc.id if item.mon_hoc else None,
            'ma_mon': item.mon_hoc.ma_mon if item.mon_hoc else '',
            'ten_mon': item.mon_hoc.ten_mon if item.mon_hoc else '',
            'giang_vien_id': item.giang_vien.id if item.giang_vien else None,
            'ma_giang_vien': item.giang_vien.ma_so if item.giang_vien else '',
            'ten_giang_vien': item.giang_vien.ho_ten if item.giang_vien else '',
            'phong_hoc_id': item.phong_hoc.id if item.phong_hoc else None,
            'ma_phong': item.phong_hoc.ma_phong if item.phong_hoc else '',
            'ngay_hoc': item.ngay_hoc.strftime('%Y-%m-%d'),
            'tiet_bat_dau': item.tiet_bat_dau,
            'tiet_ket_thuc': item.tiet_ket_thuc,
            'si_so': item.si_so,
            'ghi_chu': item.ghi_chu
        })


    page_size = request.GET.get('page_size', '10')
    try:
        page_size = int(page_size)
    except ValueError:
        page_size = 10

    paginator = Paginator(queryset, page_size)
    trang = request.GET.get('page')
    lich_hocs = paginator.get_page(trang)

    # Tên thứ trong tuần
    thu_map = {0: 'Thứ Hai', 1: 'Thứ Ba', 2: 'Thứ Tư', 3: 'Thứ Năm', 4: 'Thứ Sáu', 5: 'Thứ Bảy', 6: 'Chủ Nhật'}
    ten_thu = thu_map.get(ngay_hien_tai.weekday(), '')

    # TÙY CHỌN FILTER TRONG TEMPLATE (Cũng phải lọc theo quyền)
    ds_phong_qs = LichHoc.objects.filter(ngay_hoc=ngay_hien_tai)
    ds_lop_qs = LichHoc.objects.filter(ngay_hoc=ngay_hien_tai, lop_hoc__isnull=False)

    if request.user.la_sinh_vien:
        lop_ids_personal = DangKyHocPhan.objects.filter(sinh_vien=request.user).values_list('lop_hoc_id', flat=True)
        ds_phong_qs = ds_phong_qs.filter(lop_hoc_id__in=lop_ids_personal)
        ds_lop_qs = ds_lop_qs.filter(lop_hoc_id__in=lop_ids_personal)
    elif request.user.la_giang_vien:
        ds_phong_qs = ds_phong_qs.filter(giang_vien=request.user)
        ds_lop_qs = ds_lop_qs.filter(giang_vien=request.user)

    ds_phong = ds_phong_qs.values_list('phong_hoc_id', 'phong_hoc__ma_phong').distinct().order_by('phong_hoc__ma_phong')
    ds_lop = ds_lop_qs.values_list('lop_hoc_id', 'lop_hoc__ten_lop').distinct().order_by('lop_hoc__ten_lop')

    # Xử lý báo hỏng nhanh qua danh sách lịch
    bao_hong_id = request.GET.get('bao_hong', '')
    bao_hong_lich = None
    bao_hong_form = None
    ds_thiet_bi_phong = []

    if bao_hong_id:
        bao_hong_lich = get_object_or_404(LichHoc, pk=bao_hong_id)
        from apps.ThietBi.models import ThietBi
        from apps.ThietBi.forms import FormBaoHong
        ds_thiet_bi_phong = ThietBi.objects.filter(phong_hoc=bao_hong_lich.phong_hoc)

        if request.method == 'POST' and request.POST.get('bao_hong_submit') == '1':
            form = FormBaoHong(request.POST)
            if form.is_valid():
                bh = form.save(commit=False)
                bh.nguoi_bao = request.user
                bh.save()

                # Đánh dấu thiết bị hỏng
                bh.thiet_bi.trang_thai = 'hong'
                bh.thiet_bi.save()

                # Tạo thông báo
                from apps.ThongBao.models import ThongBao
                from apps.NguoiDung.models import NguoiDung
                ds_nguoi_nhan = NguoiDung.objects.filter(
                    vai_tro__in=[NguoiDung.VaiTro.QUAN_TRI, NguoiDung.VaiTro.GIAO_VU]
                )
                for nguoi in ds_nguoi_nhan:
                    ThongBao.objects.create(
                        tieu_de=f'Báo hỏng: {bh.thiet_bi.ten_thiet_bi}',
                        noi_dung=f'{request.user.ho_ten} báo hỏng {bh.thiet_bi.ten_thiet_bi} '
                                 f'tại phòng {bh.thiet_bi.phong_hoc.ma_phong}. Mô tả: {bh.mo_ta}',
                        loai='bao_tri',
                        nguoi_tao=request.user,
                        nguoi_nhan=nguoi,
                    )
                messages.success(request, 'Đã gửi báo hỏng thiết bị thành công.')
                # Trở lại trang danh sách, giữ nguyên tham số lọc trừ báo hỏng
                redirect_url = request.path
                query_params = request.GET.copy()
                query_params.pop('bao_hong', None)
                if query_params:
                    redirect_url += '?' + query_params.urlencode()
                return redirect(redirect_url)
            else:
                bao_hong_form = form
        else:
            bao_hong_form = FormBaoHong()
            bao_hong_form.fields['thiet_bi'].queryset = ds_thiet_bi_phong

    context = {
        'lich_hocs': lich_hocs,
        'tu_khoa': tu_khoa,
        'ngay_hien_tai': ngay_hien_tai,
        'ngay_truoc': ngay_truoc,
        'ngay_sau': ngay_sau,
        'hom_nay': hom_nay,
        'ten_thu': ten_thu,
        'trang_thai_loc': trang_thai,
        'phong_loc': phong_loc,
        'lop_loc': lop_loc,
        'buoi_loc': buoi_loc,
        'tiet_loc': tiet_loc,
        'ds_phong': ds_phong,
        'ds_lop': ds_lop,
        'tiet_choices': list(range(1, 15)),
        'khung_gio': KHUNG_GIO_TIET,
        'bao_hong_lich': bao_hong_lich,
        'bao_hong_form': bao_hong_form,
        'ds_thiet_bi_phong': ds_thiet_bi_phong,
        'week_schedules_json': json.dumps(week_data),
        'year_counts_json': json.dumps(year_counts),
        'table_schedules_json': json.dumps(table_data),
        'ngay_hien_tai_str': ngay_hien_tai.strftime('%Y-%m-%d'),
    }
    return render(request, 'LichHoc/DanhSach.html', context)



@kiem_tra_quyen_lich
def them_lich(request):
    """Thêm lịch học mới."""
    if request.method == 'POST':
        form = FormLichHoc(request.POST)
        if form.is_valid():
            lich = form.save()
            # Tạo thông báo tự động cho giảng viên
            from apps.ThongBao.models import ThongBao
            from apps.NguoiDung.models import NguoiDung
            ThongBao.objects.create(
                tieu_de=f'Lịch dạy mới: {lich.mon_hoc}',
                noi_dung=f'Bạn được phân công dạy môn {lich.mon_hoc} lớp {lich.lop_hoc.ten_lop if lich.lop_hoc else "-"} tại phòng {lich.phong_hoc.ma_phong}, '
                         f'ngày {lich.ngay_hoc}, tiết {lich.tiet_bat_dau}-{lich.tiet_ket_thuc}.',
                loai='doi_lich',
                nguoi_tao=request.user,
                nguoi_nhan=lich.giang_vien
            )
            # Tạo thông báo cho sinh viên đăng ký lớp học phần
            if lich.lop_hoc:
                sv_ids = DangKyHocPhan.objects.filter(lop_hoc=lich.lop_hoc).values_list('sinh_vien_id', flat=True)
                sinh_viens = NguoiDung.objects.filter(id__in=sv_ids)
                for sv in sinh_viens:
                    ThongBao.objects.create(
                        tieu_de=f'Lịch học mới: {lich.mon_hoc}',
                        noi_dung=f'Lớp bạn có lịch học môn {lich.mon_hoc} tại phòng {lich.phong_hoc.ma_phong}, '
                                 f'ngày {lich.ngay_hoc}, tiết {lich.tiet_bat_dau}-{lich.tiet_ket_thuc}.',
                        loai='doi_lich',
                        nguoi_tao=request.user,
                        nguoi_nhan=sv
                    )
            messages.success(request, f'Đã thêm lịch học "{lich.mon_hoc}" thành công.')
            return redirect('danh_sach_lich')
    else:
        form = FormLichHoc()
    ds_mon_hoc = MonHoc.objects.order_by('ma_mon', 'ten_mon')
    return render(request, 'LichHoc/ThemMoi.html', {'form': form, 'ds_mon_hoc': ds_mon_hoc})


@kiem_tra_quyen_lich
def chinh_sua_lich(request, pk):
    """Chỉnh sửa lịch học."""
    lich = get_object_or_404(LichHoc, pk=pk)
    if request.method == 'POST':
        form = FormLichHoc(request.POST, instance=lich)
        if form.is_valid():
            lich_moi = form.save()
            from apps.ThongBao.models import ThongBao
            from apps.NguoiDung.models import NguoiDung
            # Thông báo cho Giảng viên
            ThongBao.objects.create(
                tieu_de=f'Thay đổi lịch dạy: {lich_moi.mon_hoc}',
                noi_dung=f'Lịch dạy môn {lich_moi.mon_hoc} của bạn đã được cập nhật.',
                loai='doi_lich',
                nguoi_tao=request.user,
                nguoi_nhan=lich_moi.giang_vien
            )
            # Thông báo cho Sinh viên
            if lich_moi.lop_hoc:
                sv_ids = DangKyHocPhan.objects.filter(lop_hoc=lich_moi.lop_hoc).values_list('sinh_vien_id', flat=True)
                sinh_viens = NguoiDung.objects.filter(id__in=sv_ids)
                for sv in sinh_viens:
                    ThongBao.objects.create(
                        tieu_de=f'Thay đổi lịch học: {lich_moi.mon_hoc}',
                        noi_dung=f'Lịch học môn {lich_moi.mon_hoc} của lớp bạn đã được giáo vụ cập nhật.',
                        loai='doi_lich',
                        nguoi_tao=request.user,
                        nguoi_nhan=sv
                    )
            messages.success(request, f'Đã cập nhật lịch học "{lich_moi.mon_hoc}".')
            return redirect('danh_sach_lich')
    else:
        form = FormLichHoc(instance=lich)
    ds_mon_hoc = MonHoc.objects.order_by('ma_mon', 'ten_mon')
    return render(request, 'LichHoc/ChinhSua.html', {'form': form, 'lich': lich, 'ds_mon_hoc': ds_mon_hoc})


@kiem_tra_quyen_lich
def huy_lich(request, pk):
    """Hủy lịch học."""
    lich = get_object_or_404(LichHoc, pk=pk)
    if request.method == 'POST':
        lich.trang_thai = 'da_huy'
        lich.save()
        from apps.ThongBao.models import ThongBao
        from apps.NguoiDung.models import NguoiDung
        # Thông báo cho GV
        ThongBao.objects.create(
            tieu_de=f'Hủy lịch dạy: {lich.mon_hoc}',
            noi_dung=f'Lịch dạy môn {lich.mon_hoc} ngày {lich.ngay_hoc} đã bị hủy.',
            loai='huy_lich',
            nguoi_tao=request.user,
            nguoi_nhan=lich.giang_vien
        )
        # Thông báo cho SV
        if lich.lop_hoc:
            sv_ids = DangKyHocPhan.objects.filter(lop_hoc=lich.lop_hoc).values_list('sinh_vien_id', flat=True)
            sinh_viens = NguoiDung.objects.filter(id__in=sv_ids)
            for sv in sinh_viens:
                ThongBao.objects.create(
                    tieu_de=f'Hủy lịch học: {lich.mon_hoc}',
                    noi_dung=f'Lịch học môn {lich.mon_hoc} ngày {lich.ngay_hoc} đã bị hủy.',
                    loai='huy_lich',
                    nguoi_tao=request.user,
                    nguoi_nhan=sv
                )
        messages.success(request, f'Đã hủy lịch học "{lich.mon_hoc}".')
        return redirect('danh_sach_lich')
    return render(request, 'LichHoc/XacNhanHuy.html', {'lich': lich})


@kiem_tra_quyen_lich
def xoa_lich(request, pk):
    """Xóa vĩnh viễn lịch học."""
    lich = get_object_or_404(LichHoc, pk=pk)
    if request.method == 'POST':
        mon_hoc_ten = str(lich.mon_hoc)
        lich.delete()
        messages.success(request, f'Đã xóa vĩnh viễn lịch học môn "{mon_hoc_ten}".')
    return redirect('danh_sach_lich')


@login_required
def tra_cuu_phong(request):
    """Tra cứu phòng trống theo ngày + tiết."""
    ket_qua = None
    page_obj = None
    ngay = request.GET.get('ngay', '')
    if not ngay:
        ngay = timezone.now().date().strftime('%Y-%m-%d')
        
    tiet_bd = request.GET.get('tiet_bd', '')
    tiet_kt = request.GET.get('tiet_kt', '')
    
    if not tiet_bd or not tiet_kt:
        # Tự động xác định tiết hiện tại dựa trên giờ hệ thống
        current_hour = timezone.now().hour
        if current_hour < 12:
            tiet_bd = tiet_bd or '1'
            tiet_kt = tiet_kt or '5'
        elif current_hour < 18:
            tiet_bd = tiet_bd or '6'
            tiet_kt = tiet_kt or '10'
        else:
            tiet_bd = tiet_bd or '11'
            tiet_kt = tiet_kt or '14'

    if ngay and tiet_bd and tiet_kt:
        try:
            # Flexible date parsing
            ngay_hien_tai = datetime.strptime(ngay, '%Y-%m-%d').date()
        except ValueError:
            ngay_hien_tai = timezone.now().date()
            ngay = ngay_hien_tai.strftime('%Y-%m-%d')
            
        # Tìm phòng đã có lịch trong khoảng thời gian
        phong_da_dung = LichHoc.objects.filter(
            ngay_hoc=ngay_hien_tai,
            trang_thai='hoat_dong',
            tiet_bat_dau__lt=int(tiet_kt),
            tiet_ket_thuc__gt=int(tiet_bd),
        ).values_list('phong_hoc_id', flat=True)

        qs = PhongHoc.objects.exclude(
            id__in=phong_da_dung
        ).exclude(trang_thai='bao_tri').order_by('ma_phong')
        
        # Lọc thêm theo tòa nhà và loại phòng
        toa_nha_loc = request.GET.get('toa_nha', '')
        if toa_nha_loc:
            qs = qs.filter(toa_nha=toa_nha_loc)
            
        loai_phong_loc = request.GET.get('loai_phong', '')
        if loai_phong_loc:
            qs = qs.filter(loai_phong=loai_phong_loc)

        page_size = request.GET.get('page_size', '10')
        try:
            page_size = int(page_size)
        except ValueError:
            page_size = 10

        paginator = Paginator(qs, page_size)
        trang = request.GET.get('page')
        ket_qua = paginator.get_page(trang)
        page_obj = ket_qua

    context = {
        'ket_qua': ket_qua,
        'page_obj': page_obj,
        'ngay': ngay,
        'tiet_bd': tiet_bd,
        'tiet_kt': tiet_kt,
        'toa_nha_loc': request.GET.get('toa_nha', ''),
        'loai_phong_loc': request.GET.get('loai_phong', ''),
        'page_size': request.GET.get('page_size', '10'),
        'ds_toa_nha': PhongHoc.objects.values_list('toa_nha', flat=True).distinct().order_by('toa_nha'),
        'loai_phong_choices': PhongHoc.LoaiPhong.choices,
    }
    return render(request, 'LichHoc/TraCuu.html', context)

@login_required
def goi_y_phong_toi_uu(request):
    """
    API Gợi ý phòng thông minh dựa trên thuật toán Heuristic Scoring.
    """
    from .optimization import algorithm_room_scoring
    
    ngay_str = request.GET.get('ngay')
    try:
        tiet_bd = int(request.GET.get('tiet_bd', 1))
        tiet_kt = int(request.GET.get('tiet_kt', 1))
        si_so = int(request.GET.get('si_so', 0))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Dữ liệu đầu vào không hợp lệ'}, status=400)
        
    khoa_id = request.GET.get('khoa', '')
    lop_id = request.GET.get('lop_id')

    if not ngay_str:
        return JsonResponse({'error': 'Thiếu ngày học'}, status=400)

    # Gọi thuật toán tập trung
    ket_qua_raw = algorithm_room_scoring(ngay_str, tiet_bd, tiet_kt, si_so, khoa_id, lop_id)

    # Format lại dữ liệu cho JSON Response
    data = []
    for item in ket_qua_raw[:10]: # Lấy top 10
        phong = item['phong_obj']
        data.append({
            'id': phong.id,
            'ten_phong': f"{phong.toa_nha}.{phong.ten_phong}",
            'suc_chua': phong.suc_chua,
            'loai_phong': phong.get_loai_phong_display(),
            'ghi_chu': phong.ghi_chu,
            'score': item['score'],
            'ly_do': item['ly_do']
        })

    return JsonResponse({'data': data})


def api_loc_lop_theo_mon(request):
    """API trả về danh sách lớp khi chọn môn hoặc ngược lại."""
    mon_id = request.GET.get('mon_id')
    lop_id = request.GET.get('lop_id')
    
    if mon_id:
        lops = LopHoc.objects.filter(mon_hoc_id=mon_id).values('id', 'ten_lop')
        return JsonResponse({'lops': list(lops)})
    
    if lop_id:
        lop = LopHoc.objects.filter(id=lop_id).select_related('mon_hoc').first()
        if lop and lop.mon_hoc:
            return JsonResponse({
                'mon_id': lop.mon_hoc.id,
                'mon_ten': lop.mon_hoc.ten_mon,
                'giang_vien_id': lop.giang_vien.id if lop.giang_vien else ''
            })
            
    return JsonResponse({'error': 'Invalid request'}, status=400)

@kiem_tra_quyen_lich
def them_lich_hang_loat(request):
    """Tạo lịch học lặp theo tuần (chọn thứ, từ ngày → đến ngày)."""
    if request.method == 'POST':
        form = FormLichHoc(request.POST)
        # Lấy thông tin lặp
        ngay_bd = request.POST.get('ngay_bat_dau', '')
        ngay_kt = request.POST.get('ngay_ket_thuc', '')
        thu_chon = request.POST.getlist('thu_trong_tuan')  # ['1','3','5']
        
        if not ngay_bd or not ngay_kt or not thu_chon:
            messages.error(request, 'Vui lòng chọn đầy đủ: ngày bắt đầu, ngày kết thúc, và thứ trong tuần.')
            ds_mon_hoc = MonHoc.objects.order_by('ma_mon', 'ten_mon')
            return render(request, 'LichHoc/ThemMoiHangLoat.html', {'form': form, 'ds_mon_hoc': ds_mon_hoc})
        
        try:
            start = datetime.strptime(ngay_bd, '%Y-%m-%d').date()
            end = datetime.strptime(ngay_kt, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Ngày không hợp lệ.')
            ds_mon_hoc = MonHoc.objects.order_by('ma_mon', 'ten_mon')
            return render(request, 'LichHoc/ThemMoiHangLoat.html', {'form': form, 'ds_mon_hoc': ds_mon_hoc})
        
        thu_ints = [int(t) for t in thu_chon]  # 0=Mon, 1=Tue,...
        
        # Validate form (trừ ngay_hoc vì sẽ tự tạo)
        # Tạm set ngay_hoc để form valid
        post_data = request.POST.copy()
        post_data['ngay_hoc'] = ngay_bd
        form = FormLichHoc(post_data)
        
        if form.is_valid():
            mon_hoc = form.cleaned_data['mon_hoc']
            lop_hoc = form.cleaned_data['lop_hoc']
            giang_vien = form.cleaned_data['giang_vien']
            phong_hoc = form.cleaned_data['phong_hoc']
            tiet_bd = form.cleaned_data['tiet_bat_dau']
            tiet_kt_val = form.cleaned_data['tiet_ket_thuc']
            si_so = form.cleaned_data['si_so']
            ghi_chu = form.cleaned_data.get('ghi_chu', '')
            
            so_lich_tao = 0
            so_trung = 0
            current = start
            while current <= end:
                if current.weekday() in thu_ints:
                    # Kiểm tra trùng phòng
                    if LichHoc.kiem_tra_trung_phong(phong_hoc.id, current, tiet_bd, tiet_kt_val):
                        so_trung += 1
                        current += timedelta(days=1)
                        continue
                    # Kiểm tra trùng GV
                    if LichHoc.kiem_tra_trung_giang_vien(giang_vien.id, current, tiet_bd, tiet_kt_val):
                        so_trung += 1
                        current += timedelta(days=1)
                        continue
                    
                    LichHoc.objects.create(
                        mon_hoc=mon_hoc,
                        lop_hoc=lop_hoc,
                        ma_lop=lop_hoc.ten_lop if lop_hoc else '',
                        giang_vien=giang_vien,
                        phong_hoc=phong_hoc,
                        ngay_hoc=current,
                        tiet_bat_dau=tiet_bd,
                        tiet_ket_thuc=tiet_kt_val,
                        si_so=si_so,
                        ghi_chu=ghi_chu,
                    )
                    so_lich_tao += 1
                current += timedelta(days=1)
            
            if so_lich_tao > 0:
                messages.success(request, f'Đã tạo {so_lich_tao} lịch học thành công. {f"({so_trung} bị bỏ qua do trùng)" if so_trung else ""}')
            else:
                messages.warning(request, f'Không tạo được lịch nào. {so_trung} xung đột phát hiện.')
            return redirect('danh_sach_lich')
    else:
        form = FormLichHoc()
    
    ds_mon_hoc = MonHoc.objects.order_by('ma_mon', 'ten_mon')
    return render(request, 'LichHoc/ThemMoiHangLoat.html', {'form': form, 'ds_mon_hoc': ds_mon_hoc})


def _doc_csv_upload(csv_file):
    file_data = csv_file.read()
    encodings = ['utf-8-sig', 'utf-16', 'windows-1258', 'utf-8']
    decoded_file = None

    for encoding in encodings:
        try:
            decoded_file = file_data.decode(encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if not decoded_file:
        raise ValueError('Không thể nhận diện bảng mã file CSV. Vui lòng dùng UTF-8.')

    io_string = io.StringIO(decoded_file)
    try:
        dialect = csv.Sniffer().sniff(decoded_file[:2000], delimiters=',;')
        return list(csv.DictReader(io_string, dialect=dialect))
    except Exception:
        io_string.seek(0)
        return list(csv.DictReader(io_string))


def _lay_queryset_lop_hoc_phan(request):
    queryset = LopHoc.objects.select_related('mon_hoc').annotate(
        so_sv=Count('dang_ky_hoc_phans')
    ).order_by('ten_lop')

    tu_khoa = request.GET.get('q', '').strip()
    if tu_khoa:
        queryset = queryset.filter(
            Q(ten_lop__icontains=tu_khoa) |
            Q(mon_hoc__ten_mon__icontains=tu_khoa) |
            Q(mon_hoc__ma_mon__icontains=tu_khoa)
        )

    khoa_loc = request.GET.get('khoa', '').strip()
    if khoa_loc:
        queryset = queryset.filter(khoa=khoa_loc)

    nien_khoa_loc = request.GET.get('nien_khoa', '').strip()
    if nien_khoa_loc:
        queryset = queryset.filter(nien_khoa=nien_khoa_loc)

    context = {
        'tu_khoa': tu_khoa,
        'khoa_loc': khoa_loc,
        'nien_khoa_loc': nien_khoa_loc,
    }
    return queryset, context


@kiem_tra_quyen_lich
def lay_tien_do_nhap_lop_csv(request):
    task_id = request.GET.get('task_id')
    progress = cache.get(f'lop_import_progress_{task_id}', 0)
    status = cache.get(f'lop_import_status_{task_id}', 'Đang xử lý...')
    return JsonResponse({'progress': progress, 'status': status})


@kiem_tra_quyen_lich
def danh_sach_lop(request):
    """Danh sách lớp học."""
    from django.db.models import Count
    queryset = LopHoc.objects.annotate(
        so_sv=Count('dang_ky_hoc_phans')
    ).order_by('ten_lop')

    # Filters
    tu_khoa = request.GET.get('q', '')
    if tu_khoa:
        queryset = queryset.filter(ten_lop__icontains=tu_khoa)
    
    khoa_loc = request.GET.get('khoa', '')
    if khoa_loc:
        queryset = queryset.filter(khoa=khoa_loc)
        
    nien_khoa_loc = request.GET.get('nien_khoa', '')
    if nien_khoa_loc:
        queryset = queryset.filter(nien_khoa=nien_khoa_loc)

    # Dữ liệu cho bộ lọc
    ds_khoa = LopHoc.objects.values_list('khoa', flat=True).distinct().exclude(khoa='')
    ds_nien_khoa = LopHoc.objects.values_list('nien_khoa', flat=True).distinct().exclude(nien_khoa='')
    
    paginator = Paginator(queryset, 15)
    trang = request.GET.get('page')
    lop_hocs = paginator.get_page(trang)
    
    return render(request, 'LichHoc/DanhSachLop.html', {
        'lop_hocs': lop_hocs,
        'tu_khoa': tu_khoa,
        'khoa_loc': khoa_loc,
        'nien_khoa_loc': nien_khoa_loc,
        'ds_khoa': ds_khoa,
        'ds_nien_khoa': ds_nien_khoa,
    })


@kiem_tra_quyen_lich
def them_lop(request):
    """Thêm lớp học mới."""
    if request.method == 'POST':
        form = FormLopHoc(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Đã thêm lớp học mới thành công.')
            return redirect('danh_sach_lop')
    else:
        form = FormLopHoc()
    return render(request, 'LichHoc/ThemMoiLop.html', {'form': form})


@kiem_tra_quyen_lich
def chinh_sua_lop(request, pk):
    """Chỉnh sửa lớp học."""
    lop = get_object_or_404(LopHoc, pk=pk)
    if request.method == 'POST':
        form = FormLopHoc(request.POST, instance=lop)
        if form.is_valid():
            form.save()
            messages.success(request, f'Đã cập nhật lớp {lop.ten_lop}.')
            return redirect('danh_sach_lop')
    else:
        form = FormLopHoc(instance=lop)
    return render(request, 'LichHoc/ChinhSuaLop.html', {'form': form, 'lop': lop})


def kiem_tra_quyen_xem_lop(view_func):
    """Decorator: giáo vụ, admin, hoặc giảng viên được xem DS sinh viên."""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('dang_nhap')
        if not (request.user.la_quan_tri or request.user.la_giao_vu or request.user.la_giang_vien):
            messages.error(request, 'Bạn không có quyền xem danh sách này.')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@kiem_tra_quyen_xem_lop
def danh_sach_sv_lop(request, pk):
    """Xem danh sách sinh viên đăng ký một lớp học phần."""
    lop = get_object_or_404(LopHoc, pk=pk)
    dang_ky_list = DangKyHocPhan.objects.filter(
        lop_hoc=lop
    ).select_related('sinh_vien').order_by('sinh_vien__ho_ten')
    
    # Lịch học của lớp này
    lich_hocs = LichHoc.objects.filter(
        lop_hoc=lop, trang_thai='hoat_dong'
    ).select_related('mon_hoc', 'phong_hoc', 'giang_vien').order_by('ngay_hoc', 'tiet_bat_dau')
    
    return render(request, 'LichHoc/DanhSachSVLop.html', {
        'lop': lop,
        'dang_ky_list': dang_ky_list,
        'lich_hocs': lich_hocs,
    })


@kiem_tra_quyen_lich
def them_sv_vao_lop(request, pk):
    """Thêm sinh viên vào lớp học phần."""
    from apps.NguoiDung.models import NguoiDung
    lop = get_object_or_404(LopHoc, pk=pk)
    if request.method == 'POST':
        sv_ids = request.POST.getlist('sinh_vien_ids')
        so_them = 0
        for sv_id in sv_ids:
            _, created = DangKyHocPhan.objects.get_or_create(
                sinh_vien_id=sv_id, lop_hoc=lop
            )
            if created:
                so_them += 1
        messages.success(request, f'Đã thêm {so_them} sinh viên vào lớp {lop.ten_lop}.')
        return redirect('danh_sach_sv_lop', pk=lop.pk)
    
    # DS sinh viên chưa đăng ký lớp này
    da_dk_ids = DangKyHocPhan.objects.filter(lop_hoc=lop).values_list('sinh_vien_id', flat=True)
    sv_chua_dk = NguoiDung.objects.filter(
        vai_tro='sinh_vien', is_active=True
    ).exclude(id__in=da_dk_ids).select_related('lop_sinh_hoat').order_by('ho_ten')
    
    return render(request, 'LichHoc/ThemSVVaoLop.html', {
        'lop': lop,
        'sv_chua_dk': sv_chua_dk,
    })


@kiem_tra_quyen_lich
def xoa_sv_khoi_lop(request, pk, sv_id):
    """Xóa sinh viên khỏi lớp học phần."""
    lop = get_object_or_404(LopHoc, pk=pk)
    dk = get_object_or_404(DangKyHocPhan, lop_hoc=lop, sinh_vien_id=sv_id)
    if request.method == 'POST':
        dk.delete()
        messages.success(request, 'Đã xóa sinh viên khỏi lớp học phần.')
    return redirect('danh_sach_sv_lop', pk=lop.pk)
@kiem_tra_quyen_lich
@transaction.atomic
def nhap_lop_hoc_csv(request):
    """Nhập danh sách lớp học phần từ CSV."""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        try:
            rows = _doc_csv_upload(request.FILES['csv_file'])
            if not rows:
                messages.error(request, 'File CSV không có dữ liệu.')
                return redirect('danh_sach_lop')

            ghi_de = request.POST.get('ghi_de') == 'on'
            errors = []
            seen_lops = set()
            total_rows = len(rows)
            task_id = request.POST.get('task_id', 'manual')

            for index, row in enumerate(rows, start=2):
                ten_lop = row.get('ten_lop', '').strip().upper()
                ma_mon = row.get('ma_mon', '').strip().upper()
                ten_mon = row.get('ten_mon', '').strip()

                if not ten_lop or not ma_mon or not ten_mon:
                    errors.append(f'Dòng {index}: Thiếu một trong các cột bắt buộc ten_lop, ma_mon, ten_mon.')
                    continue

                if ten_lop in seen_lops:
                    errors.append(f"Dòng {index}: Lớp '{ten_lop}' bị trùng trong file.")
                seen_lops.add(ten_lop)

                if not ghi_de and LopHoc.objects.filter(ten_lop=ten_lop).exists():
                    errors.append(f"Dòng {index}: Lớp '{ten_lop}' đã tồn tại.")

            if errors:
                messages.error(request, 'Lỗi dữ liệu:<br>' + '<br>'.join(errors[:10]))
                return redirect('danh_sach_lop')

            count_new = 0
            count_upd = 0

            for index, row in enumerate(rows, start=1):
                percent = int((index / total_rows) * 100)
                cache.set(f'lop_import_progress_{task_id}', percent, 300)
                cache.set(f'lop_import_status_{task_id}', f"Đang xử lý lớp {index}/{total_rows}...", 300)

                ten_lop = row.get('ten_lop', '').strip().upper()
                ma_mon = row.get('ma_mon', '').strip().upper()
                ten_mon = row.get('ten_mon', '').strip()
                khoa = row.get('khoa', '').strip()
                nien_khoa = row.get('nien_khoa', '').strip()
                ten_gv = row.get('giang_vien', '').strip()

                mon_hoc, mon_created = MonHoc.objects.get_or_create(
                    ma_mon=ma_mon,
                    defaults={'ten_mon': ten_mon},
                )
                if not mon_created and mon_hoc.ten_mon != ten_mon:
                    mon_hoc.ten_mon = ten_mon
                    mon_hoc.save(update_fields=['ten_mon'])

                lop_hoc, created = LopHoc.objects.get_or_create(ten_lop=ten_lop)
                lop_hoc.mon_hoc = mon_hoc
                lop_hoc.khoa = khoa
                lop_hoc.nien_khoa = nien_khoa
                
                # Tìm giảng viên theo tên
                if ten_gv:
                    from apps.NguoiDung.models import NguoiDung
                    gv_obj = NguoiDung.objects.filter(ho_ten__iexact=ten_gv, vai_tro='giang_vien').first()
                    if gv_obj:
                        lop_hoc.giang_vien = gv_obj
                
                lop_hoc.save()

                if created:
                    count_new += 1
                else:
                    count_upd += 1

            cache.delete(f'lop_import_progress_{task_id}')
            cache.delete(f'lop_import_status_{task_id}')
            messages.success(request, f'Đã xử lý xong: Thêm mới {count_new}, Cập nhật {count_upd} lớp học phần.')
        except Exception as exc:
            messages.error(request, f'Lỗi hệ thống khi nhập CSV: {exc}')

    return redirect('danh_sach_lop')


@kiem_tra_quyen_lich
def xuat_lop_hoc_csv(request):
    """Xuất danh sách lớp học phần ra CSV."""
    queryset, _ = _lay_queryset_lop_hoc_phan(request)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_lop_hoc_phan.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Tên lớp học phần', 'Mã học phần', 'Tên môn học', 'Khoa', 'Niên khóa', 'Số sinh viên'])

    for lop in queryset:
        writer.writerow([
            lop.ten_lop,
            lop.mon_hoc.ma_mon if lop.mon_hoc else '',
            lop.mon_hoc.ten_mon if lop.mon_hoc else '',
            lop.khoa,
            lop.nien_khoa,
            lop.so_sv,
        ])

    return response


@kiem_tra_quyen_lich
def danh_sach_lop(request):
    """Danh sÃ¡ch lá»›p há»c pháº§n."""
    queryset, filter_context = _lay_queryset_lop_hoc_phan(request)

    ds_khoa = LopHoc.objects.exclude(khoa='').values_list('khoa', flat=True).distinct().order_by('khoa')
    ds_nien_khoa = LopHoc.objects.exclude(nien_khoa='').values_list('nien_khoa', flat=True).distinct().order_by('nien_khoa')
    thong_ke = {
        'tong_lop': queryset.count(),
        'tong_mon': queryset.exclude(mon_hoc__isnull=True).values('mon_hoc_id').distinct().count(),
        'tong_sinh_vien': DangKyHocPhan.objects.filter(lop_hoc_id__in=queryset.values('id')).count(),
        'tong_khoa': queryset.exclude(khoa='').values('khoa').distinct().count(),
    }

    paginator = Paginator(queryset, 15)
    trang = request.GET.get('page')
    lop_hocs = paginator.get_page(trang)

    context = {
        'lop_hocs': lop_hocs,
        'ds_khoa': ds_khoa,
        'ds_nien_khoa': ds_nien_khoa,
        'thong_ke': thong_ke,
        'mau_csv': 'ten_lop,ma_mon,ten_mon,khoa,nien_khoa',
    }
    context.update(filter_context)
    return render(request, 'LichHoc/DanhSachLop.html', context)


@kiem_tra_quyen_lich
def them_lop(request):
    """ThÃªm lá»›p há»c pháº§n má»›i."""
    if request.method == 'POST':
        form = FormLopHoc(request.POST)
        if form.is_valid():
            lop = form.save()
            messages.success(request, f'ÄÃ£ thÃªm lá»›p {lop.ten_lop} thÃ nh cÃ´ng.')
            return redirect('danh_sach_lop')
    else:
        form = FormLopHoc()
    return render(request, 'LichHoc/ThemMoiLop.html', {'form': form})


@kiem_tra_quyen_lich
def chinh_sua_lop(request, pk):
    """Chá»‰nh sá»­a lá»›p há»c pháº§n."""
    lop = get_object_or_404(LopHoc.objects.select_related('mon_hoc'), pk=pk)
    if request.method == 'POST':
        form = FormLopHoc(request.POST, instance=lop)
        if form.is_valid():
            lop = form.save()
            messages.success(request, f'ÄÃ£ cáº­p nháº­t lá»›p {lop.ten_lop}.')
            return redirect('danh_sach_lop')
    else:
        form = FormLopHoc(instance=lop)
    return render(request, 'LichHoc/ChinhSuaLop.html', {'form': form, 'lop': lop})


@login_required
def khung_gio_hoc(request):
    """Trang hiển thị khung giờ tiết học theo quy định DUT."""
    return render(request, 'LichHoc/KhungGioHoc.html', {
        'khung_gio': KHUNG_GIO_TIET,
    })


@login_required
def thoi_khoa_bieu_tuan(request):
    """Thời khóa biểu dạng lưới theo tuần."""
    hom_nay = timezone.now().date()

    ngay_chon = request.GET.get('ngay', '')
    if ngay_chon:
        try:
            ngay_goc = datetime.strptime(ngay_chon, '%Y-%m-%d').date()
        except ValueError:
            ngay_goc = hom_nay
    else:
        ngay_goc = hom_nay

    dau_tuan = ngay_goc - timedelta(days=ngay_goc.weekday())
    cuoi_tuan = dau_tuan + timedelta(days=6)
    tuan_truoc = dau_tuan - timedelta(days=7)
    tuan_sau = dau_tuan + timedelta(days=7)

    thu_map = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'CN']
    ds_ngay = []
    for i in range(7):
        ngay = dau_tuan + timedelta(days=i)
        ds_ngay.append({'ngay': ngay, 'ten_thu': thu_map[i]})

    filters = {
        'ngay_hoc__gte': dau_tuan,
        'ngay_hoc__lte': cuoi_tuan,
    }

    is_personal = request.user.la_sinh_vien or request.user.la_giang_vien

    if is_personal:
        filters['trang_thai'] = 'hoat_dong'

    if request.user.la_sinh_vien:
        lop_ids = DangKyHocPhan.objects.filter(
            sinh_vien=request.user
        ).values_list('lop_hoc_id', flat=True)
        filters['lop_hoc_id__in'] = lop_ids
    elif request.user.la_giang_vien:
        filters['giang_vien'] = request.user

    lich_hocs = LichHoc.objects.filter(**filters).select_related(
        'mon_hoc', 'phong_hoc', 'giang_vien', 'lop_hoc'
    )

    ds_tiet = list(range(1, 15))

    if is_personal:
        # ─── PERSONAL VIEW (SV/GV): rowspan merged cells ───
        start_cells = {}
        skip_cells = set()
        for lich in lich_hocs:
            wd = lich.ngay_hoc.weekday()
            start = lich.tiet_bat_dau
            end = lich.tiet_ket_thuc
            rowspan = end - start + 1
            start_cells[(wd, start)] = {'lich': lich, 'rowspan': rowspan}
            for t in range(start + 1, end + 1):
                skip_cells.add((wd, t))

        grid_rows = []
        for tiet in range(1, 15):
            cells = []
            for wd in range(7):
                key = (wd, tiet)
                if key in skip_cells:
                    cells.append({'type': 'skip'})
                elif key in start_cells:
                    info = start_cells[key]
                    cells.append({'type': 'start', 'lich': info['lich'], 'rowspan': info['rowspan']})
                else:
                    cells.append({'type': 'empty'})
            grid_rows.append({'tiet': tiet, 'cells': cells})

        template_name = 'LichHoc/ThoiKhoaBieuTuan.html'
    else:
        # ─── ADMIN/GIAO_VU VIEW: list-based management ───
        # Day tab selection
        selected_day_str = request.GET.get('day', '')
        if selected_day_str:
            try:
                selected_day = datetime.strptime(selected_day_str, '%Y-%m-%d').date()
            except ValueError:
                selected_day = hom_nay
        else:
            selected_day = hom_nay if (dau_tuan <= hom_nay <= cuoi_tuan) else dau_tuan

        # Count per day for tabs
        for d_info in ds_ngay:
            d_info['count'] = lich_hocs.filter(ngay_hoc=d_info['ngay']).count()

        # Get schedules for selected day
        day_qs = lich_hocs.filter(ngay_hoc=selected_day).order_by('tiet_bat_dau')

        # SEARCH & FILTERS
        search_mon = request.GET.get('mon', '')
        search_gv = request.GET.get('gv', '')
        search_phong = request.GET.get('phong', '')
        search_lop = request.GET.get('lop', '')
        
        if search_mon:
            day_qs = day_qs.filter(
                Q(mon_hoc__ten_mon__icontains=search_mon) |
                Q(mon_hoc__ma_mon__icontains=search_mon)
            )
        if search_gv:
            day_qs = day_qs.filter(giang_vien__ho_ten__icontains=search_gv)
        if search_phong:
            day_qs = day_qs.filter(phong_hoc__ma_phong__icontains=search_phong)
        if search_lop:
            day_qs = day_qs.filter(Q(lop_hoc__ten_lop__icontains=search_lop) | Q(ma_lop__icontains=search_lop))

        day_active = day_qs.filter(trang_thai='hoat_dong')
        day_cancelled = day_qs.filter(trang_thai='da_huy')

        # Stats
        day_stats = {
            'total': day_qs.count(),
            'rooms': day_active.values('phong_hoc_id').distinct().count(),
            'gvs': day_active.values('giang_vien_id').distinct().count(),
            'cancelled': day_cancelled.count(),
        }

        # List mode
        list_mode = request.GET.get('view', 'all')
        if list_mode not in ['all', 'rooms', 'teachers', 'cancelled']:
            list_mode = 'all'

        if list_mode == 'rooms':
            room_rows = []
            for row in day_active.values('phong_hoc__ma_phong').annotate(so_lich=Count('id')).order_by('phong_hoc__ma_phong'):
                room_rows.append({
                    'ma_phong': row['phong_hoc__ma_phong'],
                    'so_lich': row['so_lich'],
                })
            list_items = room_rows
        elif list_mode == 'teachers':
            teacher_rows = []
            for row in day_active.values('giang_vien__ho_ten').annotate(so_lich=Count('id')).order_by('giang_vien__ho_ten'):
                teacher_rows.append({
                    'ho_ten': row['giang_vien__ho_ten'],
                    'so_lich': row['so_lich'],
                })
            list_items = teacher_rows
        elif list_mode == 'cancelled':
            list_items = day_cancelled
        else:
            list_items = day_qs

        paginator = Paginator(list_items, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        cho_duyet = YeuCauDoiLich.objects.filter(trang_thai='cho_duyet').count()

        grid_rows = []
        template_name = 'LichHoc/ThoiKhoaBieuTuanAdmin.html'

    context = {
        'ds_ngay': ds_ngay, 'ds_tiet': ds_tiet, 'grid_rows': grid_rows,
        'khung_gio': KHUNG_GIO_TIET, 'dau_tuan': dau_tuan, 'cuoi_tuan': cuoi_tuan,
        'tuan_truoc': tuan_truoc, 'tuan_sau': tuan_sau, 'hom_nay': hom_nay,
        'is_personal': is_personal, 'row_h': 48,
        # Admin-only
        'selected_day': locals().get('selected_day'),
        'page_obj': locals().get('page_obj'),
        'list_mode': locals().get('list_mode', 'all'),
        'day_stats': locals().get('day_stats', {}),
        'cho_duyet': locals().get('cho_duyet', 0),
        'search_mon': locals().get('search_mon', ''),
        'search_gv': locals().get('search_gv', ''),
        'search_phong': locals().get('search_phong', ''),
        'search_lop': locals().get('search_lop', ''),
    }
    return render(request, template_name, context)


# ══════════════════════════════════════════════════════════
# YÊU CẦU ĐỔI LỊCH
# ══════════════════════════════════════════════════════════

@login_required
def tao_yeu_cau_doi_lich(request, lich_pk):
    """GV tạo yêu cầu đổi lịch."""
    lich = get_object_or_404(LichHoc, pk=lich_pk)
    if not request.user.la_giang_vien:
        messages.error(request, 'Chỉ giảng viên mới có quyền gửi yêu cầu đổi lịch.')
        return redirect('danh_sach_lich')
    if lich.giang_vien != request.user:
        messages.error(request, 'Bạn chỉ có thể yêu cầu đổi lịch dạy của mình.')
        return redirect('danh_sach_lich')

    if request.method == 'POST':
        loai = request.POST.get('loai_yeu_cau', '')
        ly_do = request.POST.get('ly_do', '')
        phong_moi_id = request.POST.get('phong_moi', '')
        ngay_moi = request.POST.get('ngay_moi', '')
        tiet_bd = request.POST.get('tiet_moi_bat_dau', '')
        tiet_kt = request.POST.get('tiet_moi_ket_thuc', '')

        if not loai or not ly_do:
            messages.error(request, 'Vui lòng chọn loại yêu cầu và nhập lý do.')
            return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)

        yc = YeuCauDoiLich(
            lich_hoc=lich, nguoi_yeu_cau=request.user,
            loai_yeu_cau=loai, ly_do=ly_do,
        )
        # Validate phòng mới nếu đổi phòng
        if loai == 'doi_phong':
            if not phong_moi_id:
                messages.error(request, 'Vui lòng chọn phòng học mới.')
                return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)
            phong = PhongHoc.objects.filter(pk=phong_moi_id).first()
            if phong:
                trung = LichHoc.kiem_tra_trung_phong(
                    phong.id, lich.ngay_hoc, lich.tiet_bat_dau, lich.tiet_ket_thuc, lich.pk
                )
                if trung:
                    messages.error(request, f'Phòng {phong.ma_phong} đã có lịch trùng. Vui lòng chọn phòng khác.')
                    return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)
                yc.phong_moi = phong
        # Thông tin đổi giờ
        elif loai == 'doi_gio':
            if not ngay_moi or not tiet_bd or not tiet_kt:
                messages.error(request, 'Vui lòng điền đầy đủ ngày mới, tiết bắt đầu và tiết kết thúc mới.')
                return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)
            try:
                yc.ngay_moi = datetime.strptime(ngay_moi, '%Y-%m-%d').date()
            except ValueError:
                messages.error(request, 'Ngày mới không hợp lệ.')
                return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)
            
            try:
                yc.tiet_moi_bat_dau = int(tiet_bd)
                yc.tiet_moi_ket_thuc = int(tiet_kt)
            except ValueError:
                messages.error(request, 'Tiết học mới phải là số nguyên.')
                return redirect('tao_yeu_cau_doi_lich', lich_pk=lich.pk)

            # Validate
            if yc.ngay_moi and yc.tiet_moi_bat_dau and yc.tiet_moi_ket_thuc:
                trung_phong = LichHoc.kiem_tra_trung_phong(
                    lich.phong_hoc_id, yc.ngay_moi, yc.tiet_moi_bat_dau, yc.tiet_moi_ket_thuc, lich.pk
                )
                trung_gv = LichHoc.kiem_tra_trung_giang_vien(
                    lich.giang_vien_id, yc.ngay_moi, yc.tiet_moi_bat_dau, yc.tiet_moi_ket_thuc, lich.pk
                )
                if trung_phong:
                    messages.warning(request, 'Lưu ý: Phòng hiện tại bị trùng ở thời gian mới. Giáo vụ sẽ xem xét.')
                if trung_gv:
                    messages.warning(request, 'Lưu ý: Bạn có lịch trùng ở thời gian mới.')

        yc.save()
        messages.success(request, 'Đã gửi yêu cầu đổi lịch. Vui lòng chờ giáo vụ duyệt.')
        return redirect('lich_su_thay_doi')

    # GET: Hiển thị form
    phong_trong = PhongHoc.objects.exclude(trang_thai='bao_tri').order_by('ma_phong')
    return render(request, 'LichHoc/TaoYeuCauDoiLich.html', {
        'lich': lich, 'phong_trong': phong_trong,
    })


@login_required
def yeu_cau_hoan_lich(request, pk):
    """Giảng viên gửi yêu cầu hoàn lại lịch đã hủy."""
    lich = get_object_or_404(LichHoc, pk=pk)
    
    if not request.user.la_giang_vien or lich.giang_vien != request.user:
        messages.error(request, 'Bạn không có quyền thực hiện thao tác này.')
        return redirect('danh_sach_lich')
    
    if lich.trang_thai != 'da_huy':
        messages.warning(request, 'Lịch học này không ở trạng thái đã hủy.')
        return redirect('danh_sach_lich')
        
    # Kiểm tra xem đã có yêu cầu tương tự đang chờ duyệt chưa
    exists = YeuCauDoiLich.objects.filter(lich_hoc=lich, loai_yeu_cau='hoan_lich', trang_thai='cho_duyet').exists()
    if exists:
        messages.info(request, 'Yêu cầu hoàn lại lịch này đang chờ giáo vụ duyệt.')
        return redirect('danh_sach_lich')

    if request.method == 'POST':
        ly_do = request.POST.get('ly_do', '')
        if not ly_do:
            messages.error(request, 'Vui lòng nhập lý do khôi phục lịch.')
        else:
            YeuCauDoiLich.objects.create(
                lich_hoc=lich,
                nguoi_yeu_cau=request.user,
                loai_yeu_cau='hoan_lich',
                ly_do=ly_do
            )
            messages.success(request, 'Đã gửi yêu cầu hoàn lại lịch học. Vui lòng chờ giáo vụ duyệt.')
            return redirect('danh_sach_lich')
            
    return render(request, 'LichHoc/XacNhanHoanLich.html', {'lich': lich})


@login_required
def danh_sach_yeu_cau(request):
    """DS yêu cầu đổi lịch — Admin/GV duyệt, GV xem của mình."""
    # 1. Xác định ngày lọc
    hom_nay = timezone.localdate()
    ngay_loc = request.GET.get('ngay', '')
    if ngay_loc:
        try:
            ngay_hien_tai = datetime.strptime(ngay_loc, '%Y-%m-%d').date()
        except ValueError:
            ngay_hien_tai = hom_nay
    else:
        ngay_hien_tai = None # Mặc định không lọc ngày nếu chưa chọn

    # 2. Khởi tạo queryset theo vai trò
    if request.user.la_quan_tri or request.user.la_giao_vu:
        queryset = YeuCauDoiLich.objects.select_related(
            'lich_hoc', 'lich_hoc__mon_hoc', 'lich_hoc__phong_hoc', 'lich_hoc__giang_vien',
            'lich_hoc__lop_hoc', 'nguoi_yeu_cau', 'phong_moi', 'nguoi_duyet'
        ).all()
    elif request.user.la_giang_vien:
        queryset = YeuCauDoiLich.objects.filter(
            nguoi_yeu_cau=request.user
        ).select_related(
            'lich_hoc', 'lich_hoc__mon_hoc', 'lich_hoc__phong_hoc', 'nguoi_duyet', 'phong_moi'
        )
    else:
        messages.error(request, 'Bạn không có quyền truy cập.')
        return redirect('dashboard')

    # 3. Filters
    tu_khoa = request.GET.get('q', '')
    if tu_khoa:
        queryset = queryset.filter(
            Q(lich_hoc__mon_hoc__ten_mon__icontains=tu_khoa) |
            Q(lich_hoc__mon_hoc__ma_mon__icontains=tu_khoa) |
            Q(nguoi_yeu_cau__ho_ten__icontains=tu_khoa)
        )

    trang_thai_loc = request.GET.get('trang_thai', '')
    if trang_thai_loc:
        queryset = queryset.filter(trang_thai=trang_thai_loc)
    
    loai_loc = request.GET.get('loai', '')
    if loai_loc:
        queryset = queryset.filter(loai_yeu_cau=loai_loc)

    if ngay_hien_tai:
        queryset = queryset.filter(lich_hoc__ngay_hoc=ngay_hien_tai)

    queryset = queryset.order_by('-ngay_tao')

    # 4. Dữ liệu cho week selector
    day_for_week = ngay_hien_tai or hom_nay
    dau_tuan = day_for_week - timedelta(days=day_for_week.weekday())
    cuoi_tuan = dau_tuan + timedelta(days=6)
    tuan_truoc = dau_tuan - timedelta(days=7)
    tuan_sau = dau_tuan + timedelta(days=7)

    thu_map = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'CN'}
    ds_ngay = []
    for offset in range(7):
        ngay = dau_tuan + timedelta(days=offset)
        ds_ngay.append({
            'ngay': ngay,
            'ten_thu': thu_map.get(ngay.weekday(), ''),
            'is_today': ngay == hom_nay,
            'is_selected': ngay == ngay_hien_tai,
        })

    paginator = Paginator(queryset, 15)
    trang = request.GET.get('page')
    yeu_caus = paginator.get_page(trang)

    # Stats
    cho_duyet_count = YeuCauDoiLich.objects.filter(trang_thai='cho_duyet').count()

    return render(request, 'LichHoc/DanhSachYeuCau.html', {
        'yeu_caus': yeu_caus, 
        'trang_thai_loc': trang_thai_loc,
        'loai_loc': loai_loc, 
        'cho_duyet': cho_duyet_count,
        'tu_khoa': tu_khoa,
        'ds_ngay': ds_ngay,
        'dau_tuan': dau_tuan,
        'cuoi_tuan': cuoi_tuan,
        'tuan_truoc': tuan_truoc,
        'tuan_sau': tuan_sau,
        'ngay_hien_tai': ngay_hien_tai,
        'hom_nay': hom_nay,
    })


@login_required
def chinh_sua_yeu_cau(request, pk):
    """Giảng viên chỉnh sửa yêu cầu của mình khi còn chờ duyệt."""
    yc = get_object_or_404(YeuCauDoiLich, pk=pk, nguoi_yeu_cau=request.user, trang_thai='cho_duyet')
    
    if request.method == 'POST':
        # Cập nhật thông tin từ form
        yc.loai_yeu_cau = request.POST.get('loai_yeu_cau', yc.loai_yeu_cau)
        
        phong_id = request.POST.get('phong_moi')
        if phong_id: yc.phong_moi_id = phong_id
            
        ngay_moi = request.POST.get('ngay_moi')
        if ngay_moi:
            yc.ngay_moi = datetime.strptime(ngay_moi, '%Y-%m-%d').date()
            
        tiet_bat_dau = request.POST.get('tiet_moi_bat_dau')
        if tiet_bat_dau: yc.tiet_moi_bat_dau = int(tiet_bat_dau)
            
        tiet_ket_thuc = request.POST.get('tiet_moi_ket_thuc')
        if tiet_ket_thuc: yc.tiet_moi_ket_thuc = int(tiet_ket_thuc)
            
        yc.ly_do = request.POST.get('ly_do', yc.ly_do)
        yc.save()
        
        messages.success(request, 'Đã cập nhật yêu cầu thành công.')
        return redirect('danh_sach_yeu_cau')

    from apps.PhongHoc.models import PhongHoc
    ds_phong = PhongHoc.objects.filter(trang_thai='trong')
    return render(request, 'LichHoc/ChinhSuaYeuCau.html', {
        'yc': yc, 
        'ds_phong': ds_phong,
        'khung_gio': KHUNG_GIO_TIET
    })


@kiem_tra_quyen_lich
def duyet_yeu_cau(request, pk):
    """Admin/GV duyệt hoặc từ chối yêu cầu đổi lịch."""
    yc = get_object_or_404(YeuCauDoiLich.objects.select_related(
        'lich_hoc', 'lich_hoc__mon_hoc', 'lich_hoc__phong_hoc', 'lich_hoc__giang_vien',
        'lich_hoc__lop_hoc', 'nguoi_yeu_cau', 'phong_moi',
    ), pk=pk)

    if request.method == 'POST':
        hanh_dong = request.POST.get('hanh_dong', '')
        ghi_chu = request.POST.get('ghi_chu_duyet', '')

        if hanh_dong == 'duyet':
            yc.trang_thai = 'da_duyet'
            yc.nguoi_duyet = request.user
            yc.ghi_chu_duyet = ghi_chu
            yc.ngay_duyet = timezone.now()
            yc.save()

            # Áp dụng thay đổi vào lịch gốc
            lich = yc.lich_hoc
            if yc.loai_yeu_cau == 'doi_phong' and yc.phong_moi:
                lich.phong_hoc = yc.phong_moi
                lich.save()
            elif yc.loai_yeu_cau == 'doi_gio':
                if yc.ngay_moi:
                    lich.ngay_hoc = yc.ngay_moi
                if yc.tiet_moi_bat_dau:
                    lich.tiet_bat_dau = yc.tiet_moi_bat_dau
                if yc.tiet_moi_ket_thuc:
                    lich.tiet_ket_thuc = yc.tiet_moi_ket_thuc
                lich.save()
            elif yc.loai_yeu_cau == 'huy_buoi':
                lich.trang_thai = 'da_huy'
                lich.save()
            elif yc.loai_yeu_cau == 'hoan_lich':
                lich.trang_thai = 'hoat_dong'
                lich.save()

            # Thông báo cho GV
            from apps.ThongBao.models import ThongBao
            ThongBao.objects.create(
                tieu_de=f'Yêu cầu đã được duyệt: {yc.get_loai_yeu_cau_display()}',
                noi_dung=f'Yêu cầu {yc.get_loai_yeu_cau_display()} cho môn {lich.mon_hoc} đã được duyệt.{" Ghi chú: " + ghi_chu if ghi_chu else ""}',
                loai='doi_lich', nguoi_tao=request.user, nguoi_nhan=yc.nguoi_yeu_cau,
            )
            messages.success(request, 'Đã duyệt yêu cầu và áp dụng thay đổi.')

        elif hanh_dong == 'tu_choi':
            yc.trang_thai = 'tu_choi'
            yc.nguoi_duyet = request.user
            yc.ghi_chu_duyet = ghi_chu
            yc.ngay_duyet = timezone.now()
            yc.save()

            from apps.ThongBao.models import ThongBao
            ThongBao.objects.create(
                tieu_de=f'Yêu cầu bị từ chối: {yc.get_loai_yeu_cau_display()}',
                noi_dung=f'Yêu cầu {yc.get_loai_yeu_cau_display()} cho môn {yc.lich_hoc.mon_hoc} đã bị từ chối.{" Lý do: " + ghi_chu if ghi_chu else ""}',
                loai='doi_lich', nguoi_tao=request.user, nguoi_nhan=yc.nguoi_yeu_cau,
            )
            messages.info(request, 'Đã từ chối yêu cầu.')

        return redirect('danh_sach_yeu_cau')

    return render(request, 'LichHoc/DuyetYeuCau.html', {'yc': yc})


@login_required
def lich_su_thay_doi(request):
    """Log lịch sử thay đổi — tất cả role có thể xem."""
    hom_nay = timezone.localdate()
    ngay_loc = request.GET.get('ngay', '')
    if ngay_loc:
        try:
            ngay_hien_tai = datetime.strptime(ngay_loc, '%Y-%m-%d').date()
        except ValueError:
            ngay_hien_tai = hom_nay
    else:
        ngay_hien_tai = hom_nay

    if request.user.la_sinh_vien:
        # SV xem log các lớp HP mình đăng ký
        lop_ids = DangKyHocPhan.objects.filter(
            sinh_vien=request.user
        ).values_list('lop_hoc_id', flat=True)
        queryset = YeuCauDoiLich.objects.filter(
            lich_hoc__lop_hoc_id__in=lop_ids
        ).exclude(trang_thai='cho_duyet')
    elif request.user.la_giang_vien:
        queryset = YeuCauDoiLich.objects.filter(nguoi_yeu_cau=request.user)
    else:
        queryset = YeuCauDoiLich.objects.all()

    # Lọc theo ngày hiện tại đang chọn
    queryset = queryset.filter(lich_hoc__ngay_hoc=ngay_hien_tai).order_by('-ngay_tao')

    # Thống kê nhanh theo trạng thái (sau khi lọc ngày)
    thong_ke = {
        'tong': queryset.count(),
        'cho_duyet': queryset.filter(trang_thai='cho_duyet').count(),
        'da_duyet': queryset.filter(trang_thai='da_duyet').count(),
        'tu_choi': queryset.filter(trang_thai='tu_choi').count(),
    }

    queryset = queryset.select_related(
        'lich_hoc', 'lich_hoc__mon_hoc', 'lich_hoc__phong_hoc', 'lich_hoc__giang_vien',
        'nguoi_yeu_cau', 'phong_moi', 'nguoi_duyet'
    )

    # Dữ liệu tuần hiện tại để hiển thị nút thứ
    dau_tuan = ngay_hien_tai - timedelta(days=ngay_hien_tai.weekday())
    cuoi_tuan = dau_tuan + timedelta(days=6)
    tuan_truoc = dau_tuan - timedelta(days=7)
    tuan_sau = dau_tuan + timedelta(days=7)

    thu_map = {0: 'Thứ 2', 1: 'Thứ 3', 2: 'Thứ 4', 3: 'Thứ 5', 4: 'Thứ 6', 5: 'Thứ 7', 6: 'CN'}
    ds_ngay = []
    for offset in range(7):
        ngay = dau_tuan + timedelta(days=offset)
        ds_ngay.append({
            'ngay': ngay,
            'ten_thu': thu_map.get(ngay.weekday(), ''),
            'is_today': ngay == hom_nay,
            'is_selected': ngay == ngay_hien_tai,
        })

    paginator = Paginator(queryset, 20)
    trang = request.GET.get('page')
    logs = paginator.get_page(trang)

    context = {
        'logs': logs,
        'thong_ke': thong_ke,
        'dau_tuan': dau_tuan,
        'cuoi_tuan': cuoi_tuan,
        'tuan_truoc': tuan_truoc,
        'tuan_sau': tuan_sau,
        'ds_ngay': ds_ngay,
        'ngay_hien_tai': ngay_hien_tai,
        'hom_nay': hom_nay,
    }
    return render(request, 'LichHoc/LichSuThayDoi.html', context)





@kiem_tra_quyen_lich
def download_file_mau_excel(request):
    """Tự động sinh và trả về tệp Excel mẫu để người dùng tải về."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mau Lịch Học"
    
    headers = [
        "Mã môn học", "Tên môn học", "Tên lớp học phần", "Mã giảng viên",
        "Ngày học", "Tiết bắt đầu", "Tiết kết thúc", "Mã phòng học",
        "Sĩ số", "Ghi chú"
    ]
    
    ws.append(headers)
    
    # Một số dữ liệu mẫu thực tế của DUT
    rows = [
        ["AC101", "Lý thuyết điều khiển tự động", "21CNCTM1", "2000118", "2026-06-15", 1, 4, "B102", 60, "Lịch học mẫu kì hè"],
        ["AR101", "Nguyên lý thiết kế kiến trúc", "21CNDK1", "2000125", "2026-06-15", 5, 8, "B103", 55, "Lịch thực hành"],
        ["AU101", "Lý thuyết ô tô", "21CNSH1", "2000110", "2026-06-16", 1, 3, "", "", "Không ghi phòng để test xếp phòng tự động"]
    ]
    
    for r in rows:
        ws.append(r)
        
    # Styling headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Apply header style
    for col_idx in range(1, 11):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Apply data styles and auto-adjust widths
    for row in range(2, 5):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=11)
            cell.border = thin_border
            if col in [5, 6, 7, 8, 9]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Write to a buffer and return
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="file_mau_lich_hoc.xlsx"'
    return response


@kiem_tra_quyen_lich
def nhap_lich_excel(request):
    """View upload file Excel và kiểm tra dữ liệu."""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
        import json
        import os
        from django.conf import settings
        import unicodedata

        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            messages.error(request, f'Lỗi đọc tệp Excel: {exc}')
            return render(request, 'LichHoc/NhapExcel.html')

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            messages.error(request, 'Tệp Excel rỗng.')
            return render(request, 'LichHoc/NhapExcel.html')

        if not headers:
            messages.error(request, 'Không tìm thấy dòng tiêu đề.')
            return render(request, 'LichHoc/NhapExcel.html')

        # Map tiêu đề cột
        header_map = {}
        for idx, val in enumerate(headers):
            if not val:
                continue
            val_str = str(val).strip().lower()
            # Loại bỏ dấu tiếng Việt để so sánh
            val_str = ''.join(c for c in unicodedata.normalize('NFD', val_str) if unicodedata.category(c) != 'Mn')
            val_str = val_str.replace(' ', '').replace('_', '')
            
            if 'mamon' in val_str:
                header_map['ma_mon'] = idx
            elif 'tenmon' in val_str:
                header_map['ten_mon'] = idx
            elif 'tenlop' in val_str:
                header_map['ten_lop'] = idx
            elif 'magiangvien' in val_str or 'msgv' in val_str or 'mgv' in val_str:
                header_map['ma_giang_vien'] = idx
            elif 'ngayhoc' in val_str or 'ngay' in val_str:
                header_map['ngay_hoc'] = idx
            elif 'tietbatdau' in val_str or 'tietbd' in val_str or 'batdau' in val_str:
                header_map['tiet_bat_dau'] = idx
            elif 'tietkethuc' in val_str or 'tietkt' in val_str or 'kethuc' in val_str:
                header_map['tiet_ket_thuc'] = idx
            elif 'maphong' in val_str or 'phong' in val_str:
                header_map['ma_phong'] = idx
            elif 'siso' in val_str:
                header_map['si_so'] = idx
            elif 'ghichu' in val_str:
                header_map['ghi_chu'] = idx

        # Thiết lập cột mặc định nếu không khớp tiêu đề
        required_keys = ['ma_mon', 'ten_mon', 'ten_lop', 'ma_giang_vien', 'ngay_hoc', 'tiet_bat_dau', 'tiet_ket_thuc']
        fallback_cols = {
            'ma_mon': 0, 'ten_mon': 1, 'ten_lop': 2, 'ma_giang_vien': 3,
            'ngay_hoc': 4, 'tiet_bat_dau': 5, 'tiet_ket_thuc': 6,
            'ma_phong': 7, 'si_so': 8, 'ghi_chu': 9
        }
        headers_detected = len(header_map) > 0
        for k in required_keys:
            if k not in header_map:
                header_map[k] = fallback_cols[k]
        for k in ['ma_phong', 'si_so', 'ghi_chu']:
            if k not in header_map:
                if headers_detected:
                    header_map[k] = None
                else:
                    header_map[k] = fallback_cols[k]

        errors = []
        valid_rows = []
        
        # Để kiểm tra trùng lặp trong chính file Excel
        seen_gv = {}
        seen_phong = {}
        seen_lop = {}

        def check_overlap(s1, e1, s2, e2):
            return s1 <= e2 and s2 <= e1

        from apps.NguoiDung.models import NguoiDung
        
        # Đọc dữ liệu từng dòng
        for row_num, row_values in enumerate(rows_iter, start=2):
            # Bỏ qua dòng rỗng hoàn toàn
            if not any(row_values):
                continue
                
            # Kiểm tra xem các cột bắt buộc có dữ liệu không
            missing = []
            for key in required_keys:
                idx = header_map[key]
                if idx >= len(row_values) or row_values[idx] is None or str(row_values[idx]).strip() == '':
                    missing.append(headers[idx] if idx < len(headers) else key)
            
            if missing:
                errors.append(f"Dòng {row_num}: Thiếu dữ liệu bắt buộc ở các cột ({', '.join(missing)}).")
                continue

            ma_mon = str(row_values[header_map['ma_mon']]).strip()
            ten_mon = str(row_values[header_map['ten_mon']]).strip()
            ten_lop = str(row_values[header_map['ten_lop']]).strip()
            ma_gv = str(row_values[header_map['ma_giang_vien']]).strip()
            ngay_hoc_raw = row_values[header_map['ngay_hoc']]
            tiet_bd_raw = row_values[header_map['tiet_bat_dau']]
            tiet_kt_raw = row_values[header_map['tiet_ket_thuc']]
            
            ma_phong = ''
            if header_map['ma_phong'] is not None and header_map['ma_phong'] < len(row_values) and row_values[header_map['ma_phong']] is not None:
                ma_phong = str(row_values[header_map['ma_phong']]).strip()
                
            si_so_raw = 30
            if header_map['si_so'] is not None and header_map['si_so'] < len(row_values) and row_values[header_map['si_so']] is not None:
                si_so_raw = row_values[header_map['si_so']]
                
            ghi_chu = ''
            if header_map['ghi_chu'] is not None and header_map['ghi_chu'] < len(row_values) and row_values[header_map['ghi_chu']] is not None:
                ghi_chu = str(row_values[header_map['ghi_chu']]).strip()

            # 1. Parse ngày học
            ngay_hoc = None
            if isinstance(ngay_hoc_raw, datetime):
                ngay_hoc = ngay_hoc_raw.date()
            elif hasattr(ngay_hoc_raw, 'date'):
                ngay_hoc = ngay_hoc_raw
            else:
                date_str = str(ngay_hoc_raw).strip()
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d'):
                    try:
                        ngay_hoc = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
            if not ngay_hoc:
                errors.append(f"Dòng {row_num}: Định dạng ngày học '{ngay_hoc_raw}' không hợp lệ. Vui lòng ghi dạng DD/MM/YYYY hoặc YYYY-MM-DD.")
                continue

            # 2. Parse tiết học
            try:
                tiet_bd = int(float(str(tiet_bd_raw).strip()))
                tiet_kt = int(float(str(tiet_kt_raw).strip()))
            except (ValueError, TypeError):
                errors.append(f"Dòng {row_num}: Tiết học phải là số nguyên (Tiết BĐ: '{tiet_bd_raw}', Tiết KT: '{tiet_kt_raw}').")
                continue
                
            if tiet_bd < 1 or tiet_bd > 14 or tiet_kt < 1 or tiet_kt > 14 or tiet_bd > tiet_kt:
                errors.append(f"Dòng {row_num}: Khoảng tiết học ({tiet_bd} - {tiet_kt}) không hợp lệ (tiết phải từ 1 đến 14 và Tiết bắt đầu <= Tiết kết thúc).")
                continue

            # 3. Parse sĩ số
            try:
                si_so = int(float(str(si_so_raw).strip()))
                if si_so <= 0:
                    si_so = 30
            except (ValueError, TypeError):
                si_so = 30

            # 4. Kiểm tra sự tồn tại trong CSDL
            mon_hoc = MonHoc.objects.filter(ma_mon=ma_mon).first()
            if not mon_hoc:
                errors.append(f"Dòng {row_num}: Môn học có mã '{ma_mon}' không tồn tại trong hệ thống.")
                continue

            lop_hoc = LopHoc.objects.filter(ten_lop=ten_lop).first()
            if not lop_hoc:
                errors.append(f"Dòng {row_num}: Lớp học phần '{ten_lop}' không tồn tại trong hệ thống.")
                continue

            giang_vien = NguoiDung.objects.filter(Q(ma_so=ma_gv) | Q(username=ma_gv), vai_tro='giang_vien').first()
            if not giang_vien:
                errors.append(f"Dòng {row_num}: Giảng viên mã/tên đăng nhập '{ma_gv}' không tồn tại trong hệ thống.")
                continue

            phong_hoc = None
            if ma_phong:
                phong_hoc = PhongHoc.objects.filter(ma_phong=ma_phong).first()
                if not phong_hoc:
                    errors.append(f"Dòng {row_num}: Phòng học '{ma_phong}' không tồn tại trong hệ thống.")
                    continue
                if phong_hoc.trang_thai == 'bao_tri':
                    errors.append(f"Dòng {row_num}: Phòng học '{ma_phong}' đang trong trạng thái bảo trì, không thể xếp lịch.")
                    continue
                # Kiểm tra phòng học thuộc Khu A hành chính
                if phong_hoc.ma_phong.upper().startswith('A') or 'Tòa A' in phong_hoc.toa_nha or 'Khu A' in phong_hoc.toa_nha:
                    errors.append(f"Dòng {row_num}: Phòng học '{ma_phong}' thuộc khu A (khu hành chính) nên không được xếp phòng học.")
                    continue
            else:
                # Tự động xếp phòng học tối ưu dựa trên Heuristic Scoring (sức chứa, vị trí gần khoa, v.v.)
                # Loại trừ các phòng đã bận trong hệ thống và các phòng đã được gán trước đó trong chính file Excel
                from .optimization import algorithm_room_scoring
                goi_y_phongs = algorithm_room_scoring(
                    ngay=ngay_hoc,
                    tiet_bd=tiet_bd,
                    tiet_kt=tiet_kt,
                    si_so=si_so,
                    khoa_id=lop_hoc.khoa if lop_hoc else None,
                    lop_id=lop_hoc.id if lop_hoc else None
                )
                
                for item in goi_y_phongs:
                    p = item['phong_obj']
                    key_phong = (p.id, ngay_hoc)
                    conflict_internal = False
                    for prev_bd, prev_kt, _ in seen_phong.get(key_phong, []):
                        if check_overlap(tiet_bd, tiet_kt, prev_bd, prev_kt):
                            conflict_internal = True
                            break
                    if not conflict_internal:
                        phong_hoc = p
                        break
                        
                if phong_hoc:
                    ma_phong = phong_hoc.ma_phong
                else:
                    errors.append(f"Dòng {row_num}: Không tìm thấy phòng học trống khả dụng nào phù hợp với sức chứa {si_so} vào ngày {ngay_hoc.strftime('%d/%m/%Y')} tiết {tiet_bd}-{tiet_kt}.")
                    continue

            # Kiểm tra xem lịch học này đã tồn tại chính xác trên hệ thống chưa
            exact_match_db = LichHoc.objects.filter(
                mon_hoc=mon_hoc,
                lop_hoc=lop_hoc,
                giang_vien=giang_vien,
                phong_hoc=phong_hoc,
                ngay_hoc=ngay_hoc,
                tiet_bat_dau=tiet_bd,
                tiet_ket_thuc=tiet_kt,
                trang_thai='hoat_dong'
            ).exists()

            if not exact_match_db:
                # 5. Kiểm tra trùng lịch NỘI BỘ file Excel
                key_gv = (giang_vien.id, ngay_hoc)
                gv_conflict = False
                for prev_bd, prev_kt, prev_row in seen_gv.get(key_gv, []):
                    if check_overlap(tiet_bd, tiet_kt, prev_bd, prev_kt):
                        errors.append(f"Dòng {row_num}: Trùng lịch giảng viên {giang_vien.ho_ten} dạy cùng lúc với Dòng {prev_row} trong chính file Excel.")
                        gv_conflict = True
                        break
                if not gv_conflict:
                    seen_gv.setdefault(key_gv, []).append((tiet_bd, tiet_kt, row_num))

                key_phong = (phong_hoc.id, ngay_hoc)
                phong_conflict = False
                for prev_bd, prev_kt, prev_row in seen_phong.get(key_phong, []):
                    if check_overlap(tiet_bd, tiet_kt, prev_bd, prev_kt):
                        errors.append(f"Dòng {row_num}: Trùng lịch phòng học {phong_hoc.ma_phong} bị sử dụng cùng lúc với Dòng {prev_row} trong chính file Excel.")
                        phong_conflict = True
                        break
                if not phong_conflict:
                    seen_phong.setdefault(key_phong, []).append((tiet_bd, tiet_kt, row_num))

                key_lop = (lop_hoc.id, ngay_hoc)
                lop_conflict = False
                for prev_bd, prev_kt, prev_row in seen_lop.get(key_lop, []):
                    if check_overlap(tiet_bd, tiet_kt, prev_bd, prev_kt):
                        errors.append(f"Dòng {row_num}: Lớp học phần {lop_hoc.ten_lop} bị trùng lịch học cùng lúc với Dòng {prev_row} trong chính file Excel.")
                        lop_conflict = True
                        break
                if not lop_conflict:
                    seen_lop.setdefault(key_lop, []).append((tiet_bd, tiet_kt, row_num))

                if gv_conflict or phong_conflict or lop_conflict:
                    continue

                # 6. Kiểm tra trùng lịch với HỆ THỐNG
                overlap_db_gv = LichHoc.objects.filter(
                    giang_vien=giang_vien,
                    ngay_hoc=ngay_hoc,
                    trang_thai='hoat_dong',
                    tiet_bat_dau__lte=tiet_kt,
                    tiet_ket_thuc__gte=tiet_bd
                ).select_related('lop_hoc').first()
                if overlap_db_gv:
                    errors.append(f"Dòng {row_num}: Trùng lịch giảng viên {giang_vien.ho_ten} dạy lớp {overlap_db_gv.lop_hoc.ten_lop if overlap_db_gv.lop_hoc else overlap_db_gv.ma_lop} (tiết {overlap_db_gv.tiet_bat_dau}-{overlap_db_gv.tiet_ket_thuc}) đã có trên hệ thống.")
                    continue

                overlap_db_phong = LichHoc.objects.filter(
                    phong_hoc=phong_hoc,
                    ngay_hoc=ngay_hoc,
                    trang_thai='hoat_dong',
                    tiet_bat_dau__lte=tiet_kt,
                    tiet_ket_thuc__gte=tiet_bd
                ).select_related('lop_hoc').first()
                if overlap_db_phong:
                    errors.append(f"Dòng {row_num}: Trùng lịch phòng học {phong_hoc.ma_phong} bị chiếm bởi lớp {overlap_db_phong.lop_hoc.ten_lop if overlap_db_phong.lop_hoc else overlap_db_phong.ma_lop} (tiết {overlap_db_phong.tiet_bat_dau}-{overlap_db_phong.tiet_ket_thuc}) đã có trên hệ thống.")
                    continue

                overlap_db_lop = LichHoc.objects.filter(
                    lop_hoc=lop_hoc,
                    ngay_hoc=ngay_hoc,
                    trang_thai='hoat_dong',
                    tiet_bat_dau__lte=tiet_kt,
                    tiet_ket_thuc__gte=tiet_bd
                ).first()
                if overlap_db_lop:
                    errors.append(f"Dòng {row_num}: Lớp học phần {lop_hoc.ten_lop} đã có lịch học trùng ({overlap_db_lop.tiet_bat_dau}-{overlap_db_lop.tiet_ket_thuc}) đã có trên hệ thống.")
                    continue

            # Thêm dòng hợp lệ
            valid_rows.append({
                'lop_hoc_id': lop_hoc.id,
                'ten_lop': ten_lop,
                'mon_hoc_id': mon_hoc.id,
                'ma_mon': mon_hoc.ma_mon,
                'ten_mon': mon_hoc.ten_mon,
                'giang_vien_id': giang_vien.id,
                'ma_giang_vien': giang_vien.ma_so,
                'ten_giang_vien': giang_vien.ho_ten,
                'phong_hoc_id': phong_hoc.id,
                'ma_phong': ma_phong,
                'ngay_hoc': ngay_hoc.strftime('%Y-%m-%d'),
                'tiet_bat_dau': tiet_bd,
                'tiet_ket_thuc': tiet_kt,
                'si_so': si_so,
                'ghi_chu': ghi_chu
            })

        # Trả về lỗi nếu có
        if errors:
            return render(request, 'LichHoc/NhapExcel.html', {'errors': errors[:50]}) # Trả về tối đa 50 lỗi đầu tiên

        if not valid_rows:
            messages.warning(request, 'Không tìm thấy dòng dữ liệu nào hợp lệ.')
            return render(request, 'LichHoc/NhapExcel.html')

        # Lưu dữ liệu hợp lệ vào tệp JSON tạm
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_file_name = f'import_preview_{request.user.id}.json'
        temp_file_path = os.path.join(temp_dir, temp_file_name)
        
        try:
            with open(temp_file_path, 'w', encoding='utf-8') as f:
                json.dump(valid_rows, f, ensure_ascii=False, indent=4)
            request.session['excel_import_preview_file'] = temp_file_path
        except Exception as exc:
            messages.error(request, f'Lỗi tạo tệp xem trước tạm thời: {exc}')
            return render(request, 'LichHoc/NhapExcel.html')

        return redirect('nhap_lich_excel_preview')

    return render(request, 'LichHoc/NhapExcel.html')


@kiem_tra_quyen_lich
def nhap_lich_excel_preview(request):
    """View hiển thị preview tổng quan đa chiều dạng Grid (Tuần/Tháng/Năm) của toàn trường."""
    import json
    import os
    
    file_path = request.session.get('excel_import_preview_file', '')
    if not file_path or not os.path.exists(file_path):
        messages.error(request, 'Không tìm thấy dữ liệu xem trước. Vui lòng tải lên lại.')
        return redirect('nhap_lich_excel')
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            preview_data = json.load(f)
    except Exception as exc:
        messages.error(request, f'Lỗi đọc dữ liệu xem trước: {exc}')
        return redirect('nhap_lich_excel')

    # Chuyển đổi sang JSON String để truyền vào JS ở template
    preview_json_str = json.dumps(preview_data)

    context = {
        'preview_data': preview_data,
        'preview_json_str': preview_json_str,
    }
    return render(request, 'LichHoc/NhapExcelPreview.html', context)


@kiem_tra_quyen_lich
@transaction.atomic
def nhap_lich_excel_confirm(request):
    """View xác nhận lưu lịch học từ tệp Excel tạm vào CSDL chính thức."""
    import json
    import os
    
    file_path = request.session.get('excel_import_preview_file', '')
    if not file_path or not os.path.exists(file_path):
        messages.error(request, 'Không tìm thấy dữ liệu nhập để xác nhận.')
        return redirect('nhap_lich_excel')

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            preview_data = json.load(f)
            
        count_created = 0
        from apps.ThongBao.models import ThongBao
        
        for row in preview_data:
            ngay_hoc = datetime.strptime(row['ngay_hoc'], '%Y-%m-%d').date()
            
            # Check if this exact record already exists to avoid duplication
            exists = LichHoc.objects.filter(
                mon_hoc_id=row['mon_hoc_id'],
                lop_hoc_id=row['lop_hoc_id'],
                giang_vien_id=row['giang_vien_id'],
                phong_hoc_id=row['phong_hoc_id'],
                ngay_hoc=ngay_hoc,
                tiet_bat_dau=row['tiet_bat_dau'],
                tiet_ket_thuc=row['tiet_ket_thuc'],
                trang_thai='hoat_dong'
            ).exists()
            
            if not exists:
                lich = LichHoc.objects.create(
                    mon_hoc_id=row['mon_hoc_id'],
                    lop_hoc_id=row['lop_hoc_id'],
                    ma_lop=row['ten_lop'],
                    giang_vien_id=row['giang_vien_id'],
                    phong_hoc_id=row['phong_hoc_id'],
                    ngay_hoc=ngay_hoc,
                    tiet_bat_dau=row['tiet_bat_dau'],
                    tiet_ket_thuc=row['tiet_ket_thuc'],
                    si_so=row['si_so'],
                    ghi_chu=row['ghi_chu']
                )
                
                # Gửi thông báo cho giảng viên
                ThongBao.objects.create(
                    tieu_de=f'Lịch dạy mới phân công: {row["ten_mon"]}',
                    noi_dung=f'Bạn được phân công giảng dạy lớp {row["ten_lop"]} tại phòng {row["ma_phong"]}, '
                             f'ngày {row["ngay_hoc"]} (tiết {row["tiet_bat_dau"]}-{row["tiet_ket_thuc"]}).',
                    loai='doi_lich',
                    nguoi_nhan_id=row['giang_vien_id'],
                    nguoi_tao=request.user
                )
                count_created += 1

        # Xóa file tạm
        try:
            os.remove(file_path)
        except Exception:
            pass
            
        request.session.pop('excel_import_preview_file', None)
        messages.success(request, f'Đã nhập thành công {count_created} lịch học phần mới vào cơ sở dữ liệu hệ thống.')
        return redirect('danh_sach_lich')
        
    except Exception as exc:
        messages.error(request, f'Lỗi hệ thống khi ghi cơ sở dữ liệu: {exc}')
        return redirect('nhap_lich_excel')


@kiem_tra_quyen_lich
def nhap_lich_excel_cancel(request):
    """Hủy bỏ và xóa tệp dữ liệu xem trước tạm thời."""
    import os
    file_path = request.session.get('excel_import_preview_file', '')
    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception:
            pass
            
    request.session.pop('excel_import_preview_file', None)
    messages.info(request, 'Đã hủy bỏ quá trình nhập thời khóa biểu từ Excel.')
    return redirect('nhap_lich_excel')


@kiem_tra_quyen_lich
def nhap_lich_excel_export_preview(request):
    """Xuất file Excel tổng quan xem trước theo định dạng bảng biểu giống trường."""
    import json
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    file_path = request.session.get('excel_import_preview_file', '')
    if not file_path or not os.path.exists(file_path):
        messages.error(request, 'Không tìm thấy dữ liệu xem trước để xuất.')
        return redirect('nhap_lich_excel')
        
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            preview_data = json.load(f)
    except Exception as exc:
        messages.error(request, f'Lỗi đọc dữ liệu xem trước: {exc}')
        return redirect('nhap_lich_excel')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TKB Preview"
    
    headers = [
        "STT", "Mã Lớp HP", "Mã Môn", "Tên Môn", "Thứ", "Tiết", "Số Tiết", "Buổi",
        "Mã GV", "Tên Giảng Viên", "Phòng", "Sĩ Số"
    ]
    ws.append(headers)
    
    # Helper for weekday
    days = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
    
    for idx, row in enumerate(preview_data, start=1):
        # Parse date to find weekday
        date_obj = datetime.strptime(row['ngay_hoc'], '%Y-%m-%d').date()
        weekday_str = days[date_obj.weekday()]
        
        tiet_bd = row['tiet_bat_dau']
        tiet_kt = row['tiet_ket_thuc']
        so_tiet = tiet_kt - tiet_bd + 1
        
        buoi = "Sáng" if tiet_bd <= 5 else ("Chiều" if tiet_bd <= 10 else "Tối")
        
        ws.append([
            idx,
            row['ten_lop'],
            row.get('ma_mon', ''),
            row['ten_mon'],
            weekday_str,
            tiet_bd,
            so_tiet,
            buoi,
            row.get('ma_giang_vien', ''),
            row['ten_giang_vien'],
            row['ma_phong'],
            row['si_so']
        ])
        
    # Styling headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Apply style to headers
    for col_idx in range(1, 13):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Apply style to data rows
    for r_idx in range(2, len(preview_data) + 2):
        for c_idx in range(1, 13):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            if c_idx in [1, 5, 6, 7, 8, 9, 11, 12]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="TKB_Preview_Export.xlsx"'
    return response


@kiem_tra_quyen_lich
def export_lich_excel(request):
    """Xuất danh sách lịch học (theo bộ lọc hiện tại hoặc toàn bộ) ra Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # 1. Khởi tạo queryset giống danh_sach_lich
    queryset = LichHoc.objects.select_related(
        'mon_hoc', 'giang_vien', 'phong_hoc', 'lop_hoc'
    )
    
    # Check if they want to export ALL records (e.g. all=1 in URL)
    export_all = request.GET.get('all', '')
    export_table = request.GET.get('table', '')
    
    if not export_all:
        if not export_table:
            # Lọc theo ngày
            hom_nay = timezone.now().date()
            ngay_loc = request.GET.get('ngay', '')
            if ngay_loc:
                try:
                    ngay_hien_tai = datetime.strptime(ngay_loc, '%Y-%m-%d').date()
                except ValueError:
                    ngay_hien_tai = hom_nay
            else:
                ngay_hien_tai = hom_nay
            
            queryset = queryset.filter(ngay_hoc=ngay_hien_tai)

        
        # Tìm kiếm
        tu_khoa = request.GET.get('q', '')
        if tu_khoa:
            queryset = queryset.filter(
                Q(mon_hoc__ten_mon__icontains=tu_khoa) |
                Q(mon_hoc__ma_mon__icontains=tu_khoa)
            )

        # Bộ lọc nâng cao
        phong_loc = request.GET.get('phong', '')
        if phong_loc:
            if phong_loc.isdigit():
                queryset = queryset.filter(phong_hoc_id=phong_loc)
            else:
                queryset = queryset.filter(phong_hoc__ma_phong=phong_loc)

        lop_loc = request.GET.get('lop', '')
        if lop_loc:
            queryset = queryset.filter(lop_hoc_id=lop_loc)

        buoi_loc = request.GET.get('buoi', '')
        buoi_map = {
            'sang': (1, 5),
            'chieu': (6, 10),
            'toi': (11, 14),
        }
        if buoi_loc in buoi_map:
            bd, kt = buoi_map[buoi_loc]
            queryset = queryset.filter(tiet_bat_dau__lte=kt, tiet_ket_thuc__gte=bd)

        tiet_loc = request.GET.get('tiet', '')
        if tiet_loc:
            try:
                tiet_loc_int = int(tiet_loc)
                queryset = queryset.filter(tiet_bat_dau__lte=tiet_loc_int, tiet_ket_thuc__gte=tiet_loc_int)
            except ValueError:
                pass

        trang_thai = request.GET.get('trang_thai', '')
        if trang_thai:
            queryset = queryset.filter(trang_thai=trang_thai)
    else:
        # Nếu export tất cả, không lọc theo ngày
        pass

    # Sắp xếp
    queryset = queryset.annotate(
        priority=Case(
            When(trang_thai='hoat_dong', then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        )
    ).order_by('ngay_hoc', 'priority', 'tiet_bat_dau')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TKB He Thong"
    
    headers = [
        "STT", "Mã Lớp HP", "Mã Môn", "Tên Môn", "Thứ", "Tiết", "Số Tiết", "Buổi",
        "Mã GV", "Tên Giảng Viên", "Phòng", "Sĩ Số"
    ]
    ws.append(headers)
    
    # Helper for weekday
    days = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
    
    for idx, obj in enumerate(queryset, start=1):
        weekday_str = days[obj.ngay_hoc.weekday()]
        
        tiet_bd = obj.tiet_bat_dau
        tiet_kt = obj.tiet_ket_thuc
        so_tiet = tiet_kt - tiet_bd + 1
        
        buoi = "Sáng" if tiet_bd <= 5 else ("Chiều" if tiet_bd <= 10 else "Tối")
        
        ma_phong = obj.phong_hoc.ma_phong if obj.phong_hoc else ""
        ten_lop = obj.lop_hoc.ten_lop if obj.lop_hoc else obj.ma_lop
        ma_mon = obj.mon_hoc.ma_mon if obj.mon_hoc else ""
        ten_mon = obj.mon_hoc.ten_mon if obj.mon_hoc else ""
        ma_gv = obj.giang_vien.ma_so if obj.giang_vien else ""
        ten_gv = obj.giang_vien.ho_ten if obj.giang_vien else ""
        
        ws.append([
            idx,
            ten_lop,
            ma_mon,
            ten_mon,
            weekday_str,
            tiet_bd,
            so_tiet,
            buoi,
            ma_gv,
            ten_gv,
            ma_phong,
            obj.si_so
        ])
        
    # Styling headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # Apply style to headers
    for col_idx in range(1, 13):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Apply style to data rows
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, 13):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            if c_idx in [1, 5, 6, 7, 8, 9, 11, 12]:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = "TKB_Chinh_Thuc_Loc.xlsx" if not export_all else "TKB_Chinh_Thuc_ToanBo.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

