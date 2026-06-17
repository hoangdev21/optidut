import django, os, sys
from datetime import datetime, timedelta

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'optidut.settings')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
django.setup()

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apps.LichHoc.models import MonHoc, LopHoc, LichHoc
from apps.NguoiDung.models import NguoiDung
from apps.PhongHoc.models import PhongHoc

mon_hocs = list(MonHoc.objects.all().order_by('ma_mon')[:20])
lop_hocs = list(LopHoc.objects.all().order_by('ten_lop')[:20])
giang_viens = list(NguoiDung.objects.filter(vai_tro='giang_vien', is_active=True).order_by('ma_so')[:20])
phongs = list(PhongHoc.objects.filter(trang_thai='trong').order_by('ma_phong')[:20])

if not mon_hocs or not lop_hocs or not giang_viens:
    print('Thieu du lieu trong DB!')
    sys.exit(1)

# ─── Xây dựng conflict map từ lịch đã có ──────────────────────────
# conflict: (date, phong_id, period) set — mỗi tiết chỉ 1 phòng
# teacher_busy: (date, gv_id, period) set — mỗi tiết 1 GV chỉ 1 lớp
NGAY_BAT_DAU = datetime.now().date() + timedelta(days=1)
SO_NGAY_TOI_DA = 60

existing = LichHoc.objects.filter(
    ngay_hoc__gte=NGAY_BAT_DAU,
    ngay_hoc__lt=NGAY_BAT_DAU + timedelta(days=SO_NGAY_TOI_DA),
    trang_thai='hoat_dong',
).values_list('ngay_hoc', 'phong_hoc_id', 'giang_vien_id', 'tiet_bat_dau', 'tiet_ket_thuc')

room_busy = set()
teacher_busy = set()
for ngay, ph_id, gv_id, tbd, tkt in existing:
    for t in range(tbd, tkt + 1):
        room_busy.add((ngay, ph_id, t))
        teacher_busy.add((ngay, gv_id, t))

TIET_MAU = [(1, 3), (4, 6), (7, 9), (1, 4), (5, 8), (9, 11), (1, 2), (3, 5), (6, 8), (7, 10)]

def tim_slot(ngay, phong_list, gv_id):
    """Tìm (phong, tiet_bd, tiet_kt) không conflict. Trả về None nếu không có."""
    for tbd, tkt in TIET_MAU:
        gv_free = all((ngay, gv_id, t) not in teacher_busy for t in range(tbd, tkt + 1))
        if not gv_free:
            continue
        for ph in phong_list:
            free = all((ngay, ph.id, t) not in room_busy for t in range(tbd, tkt + 1))
            if free:
                return ph, tbd, tkt
    return None

# ─── Sinh lịch ──────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mau Lich Hoc"

headers = [
    "Mã môn học", "Tên môn học", "Tên lớp học phần", "Mã giảng viên",
    "Ngày học", "Tiết bắt đầu", "Tiết kết thúc", "Mã phòng học",
    "Sĩ số", "Ghi chú"
]
ws.append(headers)

rows_written = 0

for i, mh in enumerate(mon_hocs):
    if i >= len(lop_hocs):
        break
    lop = lop_hocs[i]
    gv = giang_viens[i % len(giang_viens)]

    if not phongs:
        continue

    found = False
    ngay = NGAY_BAT_DAU
    for _ in range(SO_NGAY_TOI_DA):
        slot = tim_slot(ngay, phongs, gv.id)
        if slot:
            phong, tbd, tkt = slot
            for t in range(tbd, tkt + 1):
                room_busy.add((ngay, phong.id, t))
                teacher_busy.add((ngay, gv.id, t))

            si_so = 30
            try:
                si_so = lop.dang_ky_hoc_phans.count() or 30
            except Exception:
                pass

            ws.append([
                mh.ma_mon, mh.ten_mon, lop.ten_lop, gv.ma_so,
                ngay.strftime('%Y-%m-%d'), tbd, tkt,
                phong.ma_phong, si_so, f'Lịch mẫu {mh.ma_mon}',
            ])
            rows_written += 1
            found = True
            break
        ngay += timedelta(days=1)

    if not found:
        print(f'  [WARN] Không tìm được slot cho {mh.ma_mon} ({mh.ten_mon})')

# Định dạng
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 32
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 30

filename = 'mau_import_lich_hoc_moi.xlsx'
wb.save(filename)
print(f'Da tao file "{filename}" voi {rows_written}/{len(mon_hocs)} dong lich mau (khong conflict voi lich cu).')
print(f'  Mon hoc:    {len(mon_hocs)}')
print(f'  Lop HP:     {len(lop_hocs)}')
print(f'  Giang vien: {len(giang_viens)}')
print(f'  Phong:      {len(phongs)}')
